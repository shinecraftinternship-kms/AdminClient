import logging
import socket
import threading
from datetime import datetime
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import models
from .models import Client, ScanResult, AddonDevice, ActivityLog, ClientGroup, Setting, AdministratorProfile, LoginHistory, AuditLog, Company
from .models import Location, Department, Employee, EmployeeAssetAssignment, OrgAuditLog
from .models import AssetCategory, AssetVendor, Asset, AssetAssignment, AssetTransfer, AssetHistory, AssetDocument
from .diff_utils import compute_scan_diff
from .serializers import (
    AdminUserCreateSerializer, ProfileUpdateSerializer,
    LoginHistorySerializer, AuditLogSerializer,
)


def _client_ip(request):
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")


def get_user_company(request):
    """Get the Company for the currently authenticated user."""
    if not request.user or not request.user.is_authenticated:
        return None
    profile, _ = AdministratorProfile.objects.get_or_create(user=request.user)
    return profile.company
from .serializers import (
    ClientListSerializer, ClientDetailSerializer,
    ManualUpdateSerializer, ScanConfigSerializer,
    RegisterRequestSerializer, ApproveRequestSerializer, ApproveMultipleSerializer,
    PingRequestSerializer, ScanSubmitSerializer,
    AddonDeviceSerializer, ActivityLogSerializer,
    ClientGroupSerializer, SettingSerializer,
    LocationSerializer, LocationListSerializer,
    DepartmentSerializer, DepartmentListSerializer,
    EmployeeSerializer, EmployeeListSerializer,
    EmployeeAssetAssignmentSerializer, OrgAuditLogSerializer,
    AssetCategorySerializer, AssetCategoryListSerializer,
    AssetVendorSerializer, AssetSerializer, AssetListSerializer,
    AssetAssignmentSerializerV2, AssetTransferSerializer,
    AssetHistorySerializer, AssetDocumentSerializer,
)

logger = logging.getLogger("scanner_api")


@method_decorator(csrf_exempt, name="dispatch")
class RegisterClientView(APIView):
    def post(self, request):
        serializer = RegisterRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        key = data["registration_key"]
        hostname = data.get("hostname", "")
        platform_name = data.get("platform", "")
        client_version = data.get("client_version", "")
        fingerprint = data.get("device_fingerprint", "")

        existing = Client.objects.filter(registration_key=key).first()
        if existing:
            if existing.deleted:
                return Response({"status": "error", "message": "Client has been removed. Reinstall required."}, status=status.HTTP_403_FORBIDDEN)
            existing.hostname = hostname
            existing.platform = platform_name
            existing.client_version = client_version
            existing.last_seen = timezone.now()
            existing.last_ip = _client_ip(request)
            if fingerprint:
                existing.device_fingerprint = fingerprint
            existing.save(update_fields=["hostname", "platform", "client_version", "last_seen", "last_ip", "device_fingerprint"])
            return Response({"status": "pending", "approved": existing.approved})

        if fingerprint:
            same_device = Client.objects.filter(device_fingerprint=fingerprint, deleted=False).first()
            if same_device:
                same_device.registration_key = key
                same_device.hostname = hostname
                same_device.platform = platform_name
                same_device.client_version = client_version
                same_device.last_seen = timezone.now()
                same_device.last_ip = _client_ip(request)
                same_device.status = "online"
                same_device.save(update_fields=["registration_key", "hostname", "platform", "client_version", "last_seen", "last_ip", "status"])
                ActivityLog.objects.create(action="register", company=same_device.company, details=f"Client {hostname} re-registered (same device, new key {key})")
                return Response({"status": "ok", "auto_approved": same_device.approved})

        auto_approve = Setting.get("auto_approve", "false").lower() == "true"
        admin_key = Setting.get("admin_client_key", "")
        company = None
        if admin_key:
            admin_client = Client.objects.filter(registration_key=admin_key).first()
            if admin_client:
                company = admin_client.company
        Client.objects.create(
            registration_key=key, hostname=hostname, platform=platform_name,
            client_version=client_version, device_fingerprint=fingerprint,
            status="online" if auto_approve else "pending",
            approved=auto_approve, last_seen=timezone.now(),
            last_ip=_client_ip(request),
            company=company,
        )
        ActivityLog.objects.create(action="register", company=company, details=f"Client {hostname} registered with key {key}")
        return Response({"status": "ok", "auto_approved": auto_approve}, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class ApproveClientView(APIView):
    def post(self, request):
        serializer = ApproveRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        key = serializer.validated_data["registration_key"]

        updated = Client.objects.filter(registration_key=key).update(approved=True, status="online")
        if updated:
            company = get_user_company(request)
            ActivityLog.objects.create(action="approve", company=company, details=f"Client with key {key} approved")
            return Response({"status": "ok"})
        return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)


@method_decorator(csrf_exempt, name="dispatch")
class ApproveMultipleView(APIView):
    def post(self, request):
        serializer = ApproveMultipleSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        keys = serializer.validated_data["registration_keys"]

        count = Client.objects.filter(registration_key__in=keys).update(approved=True, status="online")
        company = get_user_company(request)
        ActivityLog.objects.create(action="approve", company=company, details=f"Bulk approved {count} clients")
        return Response({"status": "ok", "count": count})


@method_decorator(csrf_exempt, name="dispatch")
class PingClientView(APIView):
    def post(self, request):
        serializer = PingRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        key = data["registration_key"]

        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        client.status = "online"
        client.last_seen = timezone.now()
        client.last_ip = _client_ip(request)
        client.hostname = data.get("hostname", client.hostname)
        if data.get("client_version"):
            client.client_version = data["client_version"]
        if data.get("device_fingerprint"):
            client.device_fingerprint = data["device_fingerprint"]

        trigger = client.scan_requested
        if trigger:
            client.scan_requested = False

        client.save(update_fields=["status", "last_seen", "last_ip", "hostname", "client_version", "device_fingerprint", "scan_requested"])

        resp = {"status": "ok"}
        if trigger:
            resp["trigger_scan"] = True
        return Response(resp)


@method_decorator(csrf_exempt, name="dispatch")
class SubmitScanView(APIView):
    def post(self, request):
        serializer = ScanSubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        key = data.pop("registration_key")
        hostname = data.pop("hostname", "")
        scan_type = data.pop("scan_type", "scheduled")

        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        extra = data.pop("_extra", {})
        scan_data = {"hostname": hostname, "scan_type": scan_type, **data, **extra}
        ScanResult.objects.create(client=client, scan_type=scan_type, scan_data=scan_data)

        client.status = "online"
        client.last_seen = timezone.now()
        client.last_ip = _client_ip(request)
        client.os_version = data.get("os_info", {}).get("version", "")
        client.cpu_model = data.get("processor", {}).get("model", "")
        client.ram_info = data.get("ram", {}).get("capacity_gb", "")
        client.save(update_fields=["status", "last_seen", "last_ip", "os_version", "cpu_model", "ram_info"])

        ActivityLog.objects.create(
            action="scan", client=client, company=client.company,
            details=f"{scan_type} scan from {hostname}"
        )
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class ClientListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = Client.objects.filter(deleted=False).select_related("group")
        if company:
            qs = qs.filter(company=company)
        serializer = ClientListSerializer(qs, many=True)
        return Response(serializer.data)


@method_decorator(csrf_exempt, name="dispatch")
class ClientStatusView(APIView):
    def get(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response({"status": "approved" if client.approved else "pending", "client_status": client.status})


@method_decorator(csrf_exempt, name="dispatch")
class ClientDetailView(APIView):
    def get(self, request, key):
        try:
            client = Client.objects.prefetch_related("scans", "addons").select_related("group").get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = ClientDetailSerializer(client)
        data = serializer.data

        scans = data.get("scans") or []
        if len(scans) >= 2:
            data["scan_changes"] = compute_scan_diff(scans[1], scans[0])
        else:
            data["scan_changes"] = []
        return Response(data)

    def delete(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
            hostname = client.hostname
            client.deleted = True
            client.save(update_fields=["deleted"])
            ActivityLog.objects.create(action="delete", company=client.company, details=f"Deleted client {hostname} ({key})")
        except Client.DoesNotExist:
            pass
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class DeleteMultipleView(APIView):
    def post(self, request):
        keys = request.data.get("registration_keys", [])
        clients = Client.objects.filter(registration_key__in=keys)
        count = clients.count()
        clients.update(deleted=True)
        company = get_user_company(request)
        ActivityLog.objects.create(action="delete", company=company, details=f"Bulk deleted {count} clients")
        return Response({"status": "ok", "count": count})


@method_decorator(csrf_exempt, name="dispatch")
class ManualUpdateView(APIView):
    def put(self, request, key):
        serializer = ManualUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = {k: v for k, v in serializer.validated_data.items() if v is not None}

        updated = Client.objects.filter(registration_key=key).update(**data)
        if updated:
            company = get_user_company(request)
            ActivityLog.objects.create(action="update", company=company, details=f"Updated fields for client {key}")
            return Response({"status": "ok"})
        return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)


@method_decorator(csrf_exempt, name="dispatch")
class AddonListView(APIView):
    def get(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response([])
        return Response(AddonDeviceSerializer(client.addons.all(), many=True).data)

    def post(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = AddonDeviceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(client=client)
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AddonDeleteView(APIView):
    def delete(self, request, key, addon_id):
        deleted, _ = AddonDevice.objects.filter(id=addon_id, client__registration_key=key).delete()
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class ScanConfigView(APIView):
    def get(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
            return Response({"interval_seconds": client.scan_interval, "enabled": client.scan_enabled})
        except Client.DoesNotExist:
            return Response({"interval_seconds": 3600, "enabled": True})

    def put(self, request, key):
        serializer = ScanConfigSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        updated = Client.objects.filter(registration_key=key).update(
            scan_interval=data["interval_seconds"], scan_enabled=data["enabled"],
        )
        if updated:
            return Response({"status": "ok"})
        return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)


@method_decorator(csrf_exempt, name="dispatch")
class TriggerScanView(APIView):
    def post(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        client.scan_requested = True
        client.save(update_fields=["scan_requested"])
        ActivityLog.objects.create(action="scan_request", client=client, company=client.company, details=f"Scan requested for {client.hostname}")
        return Response({"status": "ok", "message": f"Scan queued for {client.hostname}"})


@method_decorator(csrf_exempt, name="dispatch")
class ScanAllView(APIView):
    def post(self, request):
        company = get_user_company(request)
        qs = Client.objects.filter(approved=True, deleted=False)
        if company:
            qs = qs.filter(company=company)
        count = qs.update(scan_requested=True)
        ActivityLog.objects.create(action="scan_request", company=company, details=f"Scan requested for {count} clients")
        return Response({"status": "ok", "message": f"Scan queued for {count} client(s)"})


@method_decorator(csrf_exempt, name="dispatch")
class LocalScanView(APIView):
    def post(self, request):
        from .scanner import collect_all, get_hostname, detect_platform

        def run_scan():
            try:
                data = collect_all()
                hostname = get_hostname()
                platform_name, _ = detect_platform()
                scan_data = {"hostname": hostname, "platform": platform_name, "scan_timestamp": datetime.now().isoformat(), "scanned_by": "admin_local", **data}

                admin_key = Setting.get("admin_client_key", "")
                admin_client = Client.objects.filter(registration_key=admin_key).first() if admin_key else None
                ScanResult.objects.create(client=admin_client, scan_type="local", scan_data=scan_data)
                if admin_client:
                    admin_client.status = "online"
                    admin_client.last_seen = timezone.now()
                    admin_client.os_version = data.get("os_info", {}).get("version", "")
                    admin_client.cpu_model = data.get("processor", {}).get("model", "")
                    admin_client.ram_info = data.get("ram", {}).get("capacity_gb", "")
                    admin_client.save(update_fields=["status", "last_seen", "os_version", "cpu_model", "ram_info"])
                logger.info("Admin local scan completed")
            except Exception as e:
                logger.error(f"Admin local scan failed: {e}", exc_info=True)

        threading.Thread(target=run_scan, daemon=True).start()
        return Response({"status": "ok", "message": "Local scan started"})


@method_decorator(csrf_exempt, name="dispatch")
class AdminClientInfoView(APIView):
    def get(self, request):
        key = Setting.get("admin_client_key", "")
        if not key:
            return Response({"registered": False})
        client = Client.objects.filter(registration_key=key).first()
        if not client:
            return Response({"registered": False})
        return Response({"registered": True, "registration_key": client.registration_key, "hostname": client.hostname, "status": client.status})


@method_decorator(csrf_exempt, name="dispatch")
class ClientScanResultsView(APIView):
    def get(self, request, key):
        try:
            client = Client.objects.get(registration_key=key)
        except Client.DoesNotExist:
            return Response({"status": "error", "message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)
        scan = ScanResult.objects.filter(client=client).order_by("-created_at").first()
        if scan:
            from .serializers import ScanResultSerializer
            return Response(ScanResultSerializer(scan).data)
        return Response(None)


@method_decorator(csrf_exempt, name="dispatch")
class ActivityLogView(APIView):
    def get(self, request):
        limit = int(request.GET.get("limit", 50))
        company = get_user_company(request)
        logs = ActivityLog.objects.select_related("client")
        if company:
            logs = logs.filter(company=company)
        logs = logs[:limit]
        return Response(ActivityLogSerializer(logs, many=True).data)


@method_decorator(csrf_exempt, name="dispatch")
class GroupListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        groups = ClientGroup.objects.all()
        if company:
            groups = groups.filter(company=company)
        return Response(ClientGroupSerializer(groups, many=True).data)

    def post(self, request):
        name = request.data.get("name", "").strip()
        if not name:
            return Response({"status": "error", "message": "Name required"}, status=status.HTTP_400_BAD_REQUEST)
        company = get_user_company(request)
        group, created = ClientGroup.objects.get_or_create(name=name, company=company, defaults={"description": request.data.get("description", "")})
        return Response(ClientGroupSerializer(group).data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


@method_decorator(csrf_exempt, name="dispatch")
class GroupDeleteView(APIView):
    def delete(self, request, group_id):
        ClientGroup.objects.filter(id=group_id).delete()
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class SettingsView(APIView):
    def get(self, request):
        company = get_user_company(request)
        return Response({
            "auto_approve": Setting.get("auto_approve", "false", company=company).lower() == "true",
            "stale_threshold_seconds": int(Setting.get("stale_threshold_seconds", "120", company=company)),
            "scan_all_interval": int(Setting.get("scan_all_interval", "86400", company=company)),
            "admin_client_key": Setting.get("admin_client_key", "", company=company),
        })

    def put(self, request):
        company = get_user_company(request)
        data = request.data
        if "auto_approve" in data:
            Setting.set("auto_approve", str(data["auto_approve"]).lower(), company=company)
        if "stale_threshold_seconds" in data:
            Setting.set("stale_threshold_seconds", str(data["stale_threshold_seconds"]), company=company)
        if "scan_all_interval" in data:
            Setting.set("scan_all_interval", str(data["scan_all_interval"]), company=company)
        ActivityLog.objects.create(action="setting_change", company=company, details="Settings updated")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AdminUsersView(APIView):
    def get(self, request):
        from django.contrib.auth.models import User
        users = User.objects.all().values("id", "username", "email", "is_superuser", "is_active", "date_joined")
        return Response(list(users))

    def post(self, request):
        from django.contrib.auth.models import User
        from .auth_utils import log_audit_event

        serializer = AdminUserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        username = data["username"].strip()
        email = data.get("email", "")
        password = data["password"]
        is_superuser = data.get("is_superuser", False)

        if User.objects.filter(username=username).exists():
            return Response({"status": "error", "message": "Username already exists"}, status=status.HTTP_400_BAD_REQUEST)
        if email and User.objects.filter(email=email).exists():
            return Response({"status": "error", "message": "Email already in use"}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(username=username, email=email, password=password)
        user.is_superuser = is_superuser
        user.save()
        company = get_user_company(request)
        AdministratorProfile.objects.get_or_create(user=user, defaults={"company": company})
        log_audit_event(user, "login_success", request, details=f"Admin user {username} created", success=True)
        ActivityLog.objects.create(action="login", company=company, details=f"Admin user {username} created")
        return Response({"status": "ok", "user": {"id": user.id, "username": user.username, "email": user.email, "is_superuser": user.is_superuser}})


@method_decorator(csrf_exempt, name="dispatch")
class AdminUserDeleteView(APIView):
    def delete(self, request, user_id):
        from django.contrib.auth.models import User
        try:
            user = User.objects.get(id=user_id)
            username = user.username
            user.delete()
            company = get_user_company(request)
            ActivityLog.objects.create(action="delete", company=company, details=f"Admin user {username} deleted")
            return Response({"status": "ok"})
        except User.DoesNotExist:
            return Response({"status": "error", "message": "User not found"}, status=status.HTTP_404_NOT_FOUND)


@method_decorator(csrf_exempt, name="dispatch")
class AdminStatsView(APIView):
    def get(self, request):
        from django.contrib.auth.models import User
        from .models import Client, ScanResult, ActivityLog
        company = get_user_company(request)
        base = Client.objects.all()
        if company:
            base = base.filter(company=company)
        total = base.count()
        online = base.filter(deleted=False, approved=True, status__in=["online", "pending"]).count()
        pending = base.filter(deleted=False, approved=False).count()
        deleted = base.filter(deleted=True).count()
        offline = total - online - pending
        scan_base = ScanResult.objects.all()
        log_base = ActivityLog.objects.all()
        if company:
            scan_base = scan_base.filter(client__company=company)
            log_base = log_base.filter(company=company)
        return Response({
            "total_admins": User.objects.filter(is_superuser=True).count(),
            "total_clients": total,
            "total_scans": scan_base.count(),
            "total_logs": log_base.count(),
            "clients_online": online,
            "clients_pending": pending,
            "clients_offline": offline,
            "clients_deleted": deleted,
        })


@method_decorator(csrf_exempt, name="dispatch")
class ScanChangesView(APIView):
    def get(self, request):
        from .serializers import ScanResultSerializer
        company = get_user_company(request)
        changes = []
        clients = Client.objects.filter(approved=True, deleted=False, scans__isnull=False).distinct()
        if company:
            clients = clients.filter(company=company)

        for client in clients:
            scans = ScanResult.objects.filter(client=client).order_by("-created_at")[:2]
            if len(scans) < 2:
                continue
            old_data = scans[1].scan_data or {}
            new_data = scans[0].scan_data or {}
            if old_data == new_data:
                continue
            diffs = compute_scan_diff(
                {"scan_data": old_data},
                {"scan_data": new_data},
            )
            if diffs:
                changes.append({
                    "client_hostname": client.hostname,
                    "client_key": client.registration_key,
                    "client_platform": client.platform,
                    "last_scan": scans[0].created_at.isoformat(),
                    "previous_scan": scans[1].created_at.isoformat(),
                    "change_count": len(diffs),
                    "changes": diffs[:50],
                })

        changes.sort(key=lambda c: c["last_scan"], reverse=True)
        return Response(changes)


@method_decorator(csrf_exempt, name="dispatch")
class ScanHistoryView(APIView):
    def get(self, request):
        query = request.GET.get("q", "").strip().lower()
        scan_type = request.GET.get("type", "").strip().lower()
        limit = int(request.GET.get("limit", 100))

        company = get_user_company(request)
        scans = ScanResult.objects.select_related("client").all().order_by("-created_at")
        if company:
            scans = scans.filter(client__company=company)

        if query:
            scans = scans.filter(
                models.Q(client__hostname__icontains=query) |
                models.Q(client__registration_key__icontains=query) |
                models.Q(client__platform__icontains=query)
            )
        if scan_type:
            scans = scans.filter(scan_type=scan_type)

        scans = scans[:limit]
        from .serializers import ScanHistorySerializer
        return Response(ScanHistorySerializer(scans, many=True).data)


@method_decorator(csrf_exempt, name="dispatch")
class ChangePasswordView(APIView):
    def post(self, request):
        from django.contrib.auth.models import User
        from .auth_utils import log_audit_event, get_client_ip
        from .validators import validate_strong_password

        user_id = request.data.get("user_id")
        old_password = request.data.get("old_password", "")
        new_password = request.data.get("new_password", "")
        if not user_id or not old_password or not new_password:
            return Response({"status": "error", "message": "All fields required"}, status=status.HTTP_400_BAD_REQUEST)

        password_errors = validate_strong_password(new_password)
        if password_errors:
            return Response({"status": "error", "message": "; ".join(password_errors)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"status": "error", "message": "User not found"}, status=status.HTTP_404_NOT_FOUND)
        if not user.check_password(old_password):
            log_audit_event(user, "login_failure", request, details="Password change failed: wrong current password", success=False)
            return Response({"status": "error", "message": "Current password is incorrect"}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(new_password)
        user.save()
        profile, _ = AdministratorProfile.objects.get_or_create(user=user)
        profile.password_changed_at = timezone.now()
        profile.save(update_fields=["password_changed_at"])
        log_audit_event(user, "password_changed", request, details="Password changed successfully")
        company = get_user_company(request)
        ActivityLog.objects.create(action="update", company=company, details=f"Password changed for user {user.username}")
        return Response({"status": "ok"})


def get_admin_client_key():
    hostname = socket.gethostname().upper().replace("-", "").replace(".", "")[:12]
    return f"ADMIN-{hostname}"


def ensure_admin_client():
    from .scanner import get_hostname as scan_hostname, detect_platform
    key = get_admin_client_key()
    client, created = Client.objects.get_or_create(
        registration_key=key,
        defaults={"hostname": scan_hostname(), "platform": detect_platform()[0] or "Unknown", "status": "online", "approved": True, "last_seen": timezone.now()},
    )
    Setting.set("admin_client_key", key)
    return key


def admin_self_scan():
    try:
        key = Setting.get("admin_client_key", "")
        if not key:
            return
        admin_client = Client.objects.filter(registration_key=key).first()
        if not admin_client:
            return
        from .scanner import collect_all, get_hostname, detect_platform
        data = collect_all()
        hostname = get_hostname()
        platform_name, _ = detect_platform()
        scan_data = {"hostname": hostname, "platform": platform_name, "scan_timestamp": datetime.now().isoformat(), "scanned_by": "admin_local", **data}
        ScanResult.objects.create(client=admin_client, scan_type="local", scan_data=scan_data)
        admin_client.status = "online"
        admin_client.last_seen = timezone.now()
        admin_client.os_version = data.get("os_info", {}).get("version", "")
        admin_client.cpu_model = data.get("processor", {}).get("model", "")
        admin_client.ram_info = data.get("ram", {}).get("capacity_gb", "")
        admin_client.save(update_fields=["status", "last_seen", "os_version", "cpu_model", "ram_info"])
        logger.info("Admin self-scan completed")
    except Exception as e:
        logger.error(f"Admin self-scan failed: {e}", exc_info=True)


@method_decorator(csrf_exempt, name="dispatch")
class AuthLoginView(APIView):
    def post(self, request):
        from django.contrib.auth import authenticate, login
        from django.contrib.auth.models import User
        from .validators import validate_email
        from .auth_utils import (
            check_account_lock, record_login_attempt, log_audit_event,
            create_login_history, get_client_ip, parse_device_info,
        )

        identifier = request.data.get("identifier", "").strip()
        password = request.data.get("password", "")
        remember_me = request.data.get("remember_me", False)

        if not identifier or not password:
            return Response({"status": "error", "message": "Email/username and password are required"}, status=status.HTTP_400_BAD_REQUEST)

        locked, minutes_left = check_account_lock(identifier)
        if locked:
            log_audit_event(None, "account_locked", request, details=f"Login attempt on locked account: {identifier}", success=False)
            return Response({"status": "error", "message": f"Account is locked. Try again in {minutes_left} minutes", "locked": True, "minutes_left": minutes_left}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        user = None
        if validate_email(identifier):
            try:
                u = User.objects.get(email=identifier)
                user = authenticate(request, username=u.username, password=password)
            except User.DoesNotExist:
                user = None
        else:
            user = authenticate(request, username=identifier, password=password)

        ip = get_client_ip(request)

        if user is None:
            record_login_attempt(identifier, ip, False)
            log_audit_event(None, "login_failure", request, details=f"Failed login for: {identifier}", success=False)
            remaining = int(Setting.get("max_login_attempts", "5"))
            attempts = __import__("scanner_api.models", fromlist=["LoginAttempt"]).LoginAttempt.objects.filter(
                identifier=identifier, success=False,
                created_at__gte=timezone.now() - timezone.timedelta(minutes=30)
            ).count()
            remaining = max(0, remaining - attempts)
            return Response({"status": "error", "message": "Invalid credentials", "attempts_remaining": remaining}, status=status.HTTP_401_UNAUTHORIZED)

        if not user.is_active:
            return Response({"status": "error", "message": "Account is disabled"}, status=status.HTTP_403_FORBIDDEN)

        record_login_attempt(identifier, ip, True)
        log_audit_event(user, "login_success", request, details=f"Login successful for {user.username}")
        login_history = create_login_history(user, request)

        if remember_me:
            request.session.set_expiry(60 * 60 * 24 * 30)
        else:
            request.session.set_expiry(0)

        request.session["last_activity"] = timezone.now().isoformat()
        request.session["login_history_id"] = login_history.id

        login(request, user)
        _profile = AdministratorProfile.objects.filter(user=user).select_related("company").first()
        ActivityLog.objects.create(action="login", company=_profile.company if _profile else None, details=f"Admin user {user.username} logged in")

        return Response({
            "status": "ok",
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "is_superuser": user.is_superuser,
                "first_name": user.first_name,
                "last_name": user.last_name,
            }
        })


@method_decorator(csrf_exempt, name="dispatch")
class AuthLogoutView(APIView):
    def post(self, request):
        from .auth_utils import log_audit_event, close_login_history
        if request.user.is_authenticated:
            log_audit_event(request.user, "logout", request, details="User logged out")
            close_login_history(request.user)
        from django.contrib.auth import logout
        logout(request)
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AuthMeView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({"authenticated": False}, status=status.HTTP_401_UNAUTHORIZED)
        user = request.user
        profile, _ = AdministratorProfile.objects.get_or_create(user=user)
        return Response({
            "authenticated": True,
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "is_superuser": user.is_superuser,
                "is_active": user.is_active,
                "date_joined": user.date_joined.isoformat() if user.date_joined else None,
                "last_login": user.last_login.isoformat() if user.last_login else None,
            },
            "profile": {
                "phone_number": profile.phone_number,
                "profile_picture_url": profile.profile_picture_url,
                "timezone": profile.timezone,
                "currency": profile.currency,
                "date_format": profile.date_format,
                "notification_email": profile.notification_email,
                "notification_in_app": profile.notification_in_app,
                "notification_daily_summary": profile.notification_daily_summary,
                "password_changed_at": profile.password_changed_at.isoformat() if profile.password_changed_at else None,
                "mfa_enabled": profile.mfa_enabled,
            }
        })


@method_decorator(csrf_exempt, name="dispatch")
class AuthProfileView(APIView):
    def put(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        from .auth_utils import log_audit_event
        serializer = ProfileUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        user = request.user
        if "first_name" in data:
            user.first_name = data["first_name"]
        if "last_name" in data:
            user.last_name = data["last_name"]
        if "email" in data:
            user.email = data["email"]
        user.save(update_fields=["first_name", "last_name", "email"])
        profile, _ = AdministratorProfile.objects.get_or_create(user=user)
        for field in ["phone_number", "timezone", "currency", "date_format",
                       "notification_email", "notification_in_app", "notification_daily_summary",
                       "dashboard_default"]:
            if field in data:
                setattr(profile, field, data[field])
        profile.save()
        log_audit_event(user, "profile_updated", request, details="Profile updated")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AuthLoginHistoryView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        limit = int(request.GET.get("limit", 50))
        offset = int(request.GET.get("offset", 0))
        query = request.GET.get("q", "").strip().lower()

        qs = LoginHistory.objects.filter(user=request.user)
        if query:
            from django.db import models as db_models
            qs = qs.filter(
                db_models.Q(browser__icontains=query) |
                db_models.Q(os__icontains=query) |
                db_models.Q(ip_address__icontains=query)
            )
        total = qs.count()
        entries = qs[offset:offset + limit]
        return Response({
            "total": total,
            "entries": LoginHistorySerializer(entries, many=True).data,
        })


@method_decorator(csrf_exempt, name="dispatch")
class AuthAuditLogsView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        limit = int(request.GET.get("limit", 50))
        offset = int(request.GET.get("offset", 0))
        event_type = request.GET.get("event_type", "").strip()
        user_id = request.GET.get("user_id", "")

        qs = AuditLog.objects.all()
        if event_type:
            qs = qs.filter(event_type=event_type)
        if user_id:
            qs = qs.filter(user_id=user_id)
        total = qs.count()
        entries = qs[offset:offset + limit]
        return Response({
            "total": total,
            "entries": AuditLogSerializer(entries, many=True).data,
        })


@method_decorator(csrf_exempt, name="dispatch")
class AuthActiveSessionsView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        sessions = LoginHistory.objects.filter(user=request.user, is_current=True)
        return Response(LoginHistorySerializer(sessions, many=True).data)


@method_decorator(csrf_exempt, name="dispatch")
class OrgSettingsView(APIView):
    def get(self, request):
        company = get_user_company(request)
        return Response({
            "org_name": Setting.get("org_name", "IT Asset Management System", company=company),
            "org_logo_url": Setting.get("org_logo_url", "", company=company),
            "org_timezone": Setting.get("org_timezone", "UTC", company=company),
            "org_currency": Setting.get("org_currency", "USD", company=company),
            "org_date_format": Setting.get("org_date_format", "YYYY-MM-DD", company=company),
        })

    def put(self, request):
        from .auth_utils import log_audit_event
        company = get_user_company(request)
        data = request.data
        for key in ["org_name", "org_logo_url", "org_timezone", "org_currency", "org_date_format"]:
            if key in data:
                Setting.set(key, data[key], company=company)
        if request.user.is_authenticated:
            log_audit_event(request.user, "settings_updated", request, details="Organization settings updated")
        ActivityLog.objects.create(action="setting_change", company=company, details="Organization settings updated")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class SecuritySettingsView(APIView):
    def get(self, request):
        company = get_user_company(request)
        return Response({
            "session_timeout_minutes": int(Setting.get("session_timeout_minutes", "30", company=company)),
            "max_login_attempts": int(Setting.get("max_login_attempts", "5", company=company)),
            "lock_duration_minutes": int(Setting.get("lock_duration_minutes", "30", company=company)),
            "password_expiry_days": int(Setting.get("password_expiry_days", "0", company=company)),
        })

    def put(self, request):
        from .auth_utils import log_audit_event
        company = get_user_company(request)
        data = request.data
        for key in ["session_timeout_minutes", "max_login_attempts", "lock_duration_minutes", "password_expiry_days"]:
            if key in data:
                Setting.set(key, str(data[key]), company=company)
        if request.user.is_authenticated:
            log_audit_event(request.user, "settings_updated", request, details="Security settings updated")
        ActivityLog.objects.create(action="setting_change", company=company, details="Security settings updated")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class NotificationSettingsView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        profile, _ = AdministratorProfile.objects.get_or_create(user=request.user)
        return Response({
            "notification_email": profile.notification_email,
            "notification_in_app": profile.notification_in_app,
            "notification_daily_summary": profile.notification_daily_summary,
        })

    def put(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        profile, _ = AdministratorProfile.objects.get_or_create(user=request.user)
        data = request.data
        for field in ["notification_email", "notification_in_app", "notification_daily_summary"]:
            if field in data:
                setattr(profile, field, data[field])
        profile.save()
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class DashboardSettingsView(APIView):
    def get(self, request):
        return Response({
            "dashboard_default": Setting.get("dashboard_default", "dashboard"),
            "dashboard_filters": Setting.get("dashboard_filters", "{}"),
        })

    def put(self, request):
        data = request.data
        if "dashboard_default" in data:
            Setting.set("dashboard_default", data["dashboard_default"])
        if "dashboard_filters" in data:
            import json
            Setting.set("dashboard_filters", json.dumps(data["dashboard_filters"]))
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AuthAvatarUploadView(APIView):
    def post(self, request):
        if not request.user.is_authenticated:
            return Response({"status": "error", "message": "Not authenticated"}, status=status.HTTP_401_UNAUTHORIZED)
        if "avatar" not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        f = request.FILES["avatar"]
        if f.size > 2 * 1024 * 1024:
            return Response({"status": "error", "message": "File too large (max 2MB)"}, status=status.HTTP_400_BAD_REQUEST)
        if f.content_type not in ("image/jpeg", "image/png", "image/webp"):
            return Response({"status": "error", "message": "Only JPEG, PNG, and WebP are allowed"}, status=status.HTTP_400_BAD_REQUEST)

        import base64
        file_data = base64.b64encode(f.read()).decode("utf-8")
        ext = f.name.rsplit(".", 1)[-1] if "." in f.name else "jpg"
        content_type = f.content_type
        data_url = f"data:{content_type};base64,{file_data}"

        profile, _ = AdministratorProfile.objects.get_or_create(user=request.user)
        profile.profile_picture_url = data_url
        profile.save(update_fields=["profile_picture_url"])
        return Response({"status": "ok", "profile_picture_url": data_url})


# ═══════════════════════════════════════════════════════════════════════════════
# ORGANIZATION MODULE - Employee, Department, Location Management
# ═══════════════════════════════════════════════════════════════════════════════


def _org_audit(request, entity_type, entity_id, entity_name, action, prev=None, new=None, company=None):
    OrgAuditLog.objects.create(
        entity_type=entity_type,
        entity_id=str(entity_id),
        entity_name=entity_name,
        action=action,
        previous_value=prev or {},
        new_value=new or {},
        performed_by=request.user.username if request.user.is_authenticated else "system",
        ip_address=_client_ip(request),
        user_agent=request.META.get("HTTP_USER_AGENT", ""),
        company=company,
    )


# ── Location Views ──────────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class LocationListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = Location.objects.filter(deleted=False)
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        city = request.query_params.get("city", "").strip()
        s = request.query_params.get("status", "").strip()
        if search:
            qs = qs.filter(
                models.Q(office_name__icontains=search) |
                models.Q(building_name__icontains=search) |
                models.Q(city__icontains=search)
            )
        if city:
            qs = qs.filter(city__icontains=city)
        if s:
            qs = qs.filter(status=s)
        return Response(LocationListSerializer(qs, many=True).data)

    def post(self, request):
        company = get_user_company(request)
        serializer = LocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if Location.objects.filter(
            office_name=data["office_name"], city=data["city"], deleted=False, company=company
        ).exists():
            return Response(
                {"status": "error", "message": "A location with this office name and city already exists"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        loc = serializer.save(company=company)
        _org_audit(request, "location", loc.id, loc.office_name, "created", new=serializer.data)
        return Response(LocationSerializer(loc).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class LocationDetailView(APIView):
    def get(self, request, key):
        try:
            loc = Location.objects.get(id=key, deleted=False)
        except Location.DoesNotExist:
            return Response({"status": "error", "message": "Location not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(LocationSerializer(loc).data)

    def put(self, request, key):
        try:
            loc = Location.objects.get(id=key, deleted=False)
        except Location.DoesNotExist:
            return Response({"status": "error", "message": "Location not found"}, status=status.HTTP_404_NOT_FOUND)
        if loc.status == "Closed":
            return Response({"status": "error", "message": "Closed locations are read-only"}, status=status.HTTP_400_BAD_REQUEST)
        prev = LocationSerializer(loc).data
        serializer = LocationSerializer(loc, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        new_office = serializer.validated_data.get("office_name", loc.office_name)
        new_city = serializer.validated_data.get("city", loc.city)
        dup = Location.objects.filter(office_name=new_office, city=new_city, deleted=False).exclude(id=key)
        if dup.exists():
            return Response(
                {"status": "error", "message": "A location with this office name and city already exists"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        updated = serializer.save()
        _org_audit(request, "location", key, updated.office_name, "updated", prev=prev, new=serializer.data)
        return Response(LocationSerializer(updated).data)


@method_decorator(csrf_exempt, name="dispatch")
class LocationDeleteView(APIView):
    def post(self, request, key):
        try:
            loc = Location.objects.get(id=key, deleted=False)
        except Location.DoesNotExist:
            return Response({"status": "error", "message": "Location not found"}, status=status.HTTP_404_NOT_FOUND)
        active_assets = EmployeeAssetAssignment.objects.filter(
            is_active=True, employee__location=loc
        ).count()
        if active_assets > 0:
            return Response(
                {"status": "error", "message": f"Cannot delete location with {active_assets} active asset(s)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        loc.delete()
        _org_audit(request, "location", key, loc.office_name, "deleted")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class LocationArchiveView(APIView):
    def post(self, request, key):
        try:
            loc = Location.objects.get(id=key, deleted=False)
        except Location.DoesNotExist:
            return Response({"status": "error", "message": "Location not found"}, status=status.HTTP_404_NOT_FOUND)
        active_assets = EmployeeAssetAssignment.objects.filter(
            is_active=True, employee__location=loc
        ).count()
        if active_assets > 0:
            return Response(
                {"status": "error", "message": f"Cannot archive location with {active_assets} active asset(s)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        prev_status = loc.status
        loc.status = "Archived"
        loc.save(update_fields=["status"])
        _org_audit(request, "location", key, loc.office_name, "archived",
                   prev={"status": prev_status}, new={"status": "Archived"})
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class LocationDashboardView(APIView):
    def get(self, request, key):
        try:
            loc = Location.objects.get(id=key, deleted=False)
        except Location.DoesNotExist:
            return Response({"status": "error", "message": "Location not found"}, status=status.HTTP_404_NOT_FOUND)
        employees = Employee.objects.filter(location=loc, deleted=False)
        assignments = EmployeeAssetAssignment.objects.filter(is_active=True, employee__location=loc)
        total_assets = assignments.values("client").distinct().count()
        online_devices = Client.objects.filter(
            employee_assignments__in=assignments, status="online", deleted=False
        ).distinct().count()
        offline_devices = total_assets - online_devices
        asset_value = sum(
            (a.client.purchase_cost for a in assignments.select_related("client") if a.client and a.client.purchase_cost),
            __import__("decimal").Decimal("0"),
        )
        recent = ActivityLog.objects.filter(
            client__employee_assignments__employee__location=loc
        ).order_by("-created_at")[:10]
        return Response({
            "total_assets": total_assets,
            "online_devices": online_devices,
            "offline_devices": max(0, offline_devices),
            "active_employees": employees.filter(status="Active").count(),
            "asset_value": float(asset_value),
            "recent_activities": ActivityLogSerializer(recent, many=True).data,
        })


@method_decorator(csrf_exempt, name="dispatch")
class LocationExportView(APIView):
    def get(self, request):
        import csv
        from django.http import HttpResponse
        company = get_user_company(request)
        fmt = request.query_params.get("format", "csv")
        qs = Location.objects.filter(deleted=False)
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(models.Q(office_name__icontains=search) | models.Q(city__icontains=search))
        if fmt == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="locations.csv"'
            writer = csv.writer(response)
            writer.writerow(["Office Name", "Building", "Floor", "Room", "City", "State", "Country", "Status", "Manager"])
            for loc in qs:
                writer.writerow([loc.office_name, loc.building_name, loc.floor, loc.room_number,
                                 loc.city, loc.state, loc.country, loc.status, loc.office_manager])
            return response
        locs = LocationListSerializer(qs, many=True).data
        return Response(locs)


@method_decorator(csrf_exempt, name="dispatch")
class LocationImportView(APIView):
    def post(self, request):
        import csv, io
        f = request.FILES.get("file")
        if not f:
            return Response({"status": "error", "message": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = request.data.get("dry_run", "false") == "true"
        try:
            content = f.read().decode("utf-8")
        except UnicodeDecodeError:
            return Response({"status": "error", "message": "File must be UTF-8 encoded"}, status=status.HTTP_400_BAD_REQUEST)
        reader = csv.DictReader(io.StringIO(content))
        success_count = 0
        errors = []
        for i, row in enumerate(reader, start=2):
            office_name = row.get("office_name", "").strip()
            city = row.get("city", "").strip()
            if not office_name or not city:
                errors.append({"row": i, "message": "office_name and city are required"})
                continue
            if Location.objects.filter(office_name=office_name, city=city, deleted=False).exists():
                errors.append({"row": i, "message": f"Location '{office_name}' in '{city}' already exists"})
                continue
            if not dry_run:
                Location.objects.create(
                    office_name=office_name, building_name=row.get("building_name", ""),
                    floor=row.get("floor", ""), room_number=row.get("room_number", ""),
                    address=row.get("address", ""), city=city, state=row.get("state", ""),
                    country=row.get("country", "USA"), postal_code=row.get("postal_code", ""),
                    contact_number=row.get("contact_number", ""),
                    office_manager=row.get("office_manager", ""),
                    status=row.get("status", "Active"),
                )
            success_count += 1
        return Response({"status": "ok", "success_count": success_count, "errors": errors, "dry_run": dry_run})


@method_decorator(csrf_exempt, name="dispatch")
class LocationBulkActionView(APIView):
    def post(self, request):
        action = request.data.get("action", "")
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"status": "error", "message": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        qs = Location.objects.filter(id__in=ids, deleted=False)
        count = 0
        if action == "archive":
            count = qs.update(status="Archived")
        elif action == "delete":
            for loc in qs:
                active = EmployeeAssetAssignment.objects.filter(is_active=True, employee__location=loc).count()
                if active == 0:
                    loc.delete()
                    count += 1
        return Response({"status": "ok", "count": count})


# ── Department Views ────────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = Department.objects.filter(deleted=False)
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        s = request.query_params.get("status", "").strip()
        if search:
            qs = qs.filter(
                models.Q(name__icontains=search) |
                models.Q(code__icontains=search) |
                models.Q(department_head__icontains=search)
            )
        if s:
            qs = qs.filter(status=s)
        return Response(DepartmentListSerializer(qs, many=True).data)

    def post(self, request):
        company = get_user_company(request)
        serializer = DepartmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if Department.objects.filter(name=data["name"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "Department name already exists"}, status=status.HTTP_400_BAD_REQUEST)
        if Department.objects.filter(code=data["code"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "Department code already exists"}, status=status.HTTP_400_BAD_REQUEST)
        dept = serializer.save(company=company)
        _org_audit(request, "department", dept.id, dept.name, "created", new=serializer.data, company=company)
        return Response(DepartmentSerializer(dept).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentDetailView(APIView):
    def get(self, request, key):
        try:
            dept = Department.objects.get(id=key, deleted=False)
        except Department.DoesNotExist:
            return Response({"status": "error", "message": "Department not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(DepartmentSerializer(dept).data)

    def put(self, request, key):
        try:
            dept = Department.objects.get(id=key, deleted=False)
        except Department.DoesNotExist:
            return Response({"status": "error", "message": "Department not found"}, status=status.HTTP_404_NOT_FOUND)
        prev = DepartmentSerializer(dept).data
        serializer = DepartmentSerializer(dept, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        new_name = serializer.validated_data.get("name", dept.name)
        new_code = serializer.validated_data.get("code", dept.code)
        if Department.objects.filter(name=new_name, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "Department name already exists"}, status=status.HTTP_400_BAD_REQUEST)
        if Department.objects.filter(code=new_code, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "Department code already exists"}, status=status.HTTP_400_BAD_REQUEST)
        updated = serializer.save()
        _org_audit(request, "department", key, updated.name, "updated", prev=prev, new=serializer.data)
        return Response(DepartmentSerializer(updated).data)


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentDeleteView(APIView):
    def post(self, request, key):
        try:
            dept = Department.objects.get(id=key, deleted=False)
        except Department.DoesNotExist:
            return Response({"status": "error", "message": "Department not found"}, status=status.HTTP_404_NOT_FOUND)
        if Employee.objects.filter(department=dept, deleted=False).exists():
            return Response(
                {"status": "error", "message": "Cannot delete department with active employees"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        active_assets = EmployeeAssetAssignment.objects.filter(
            is_active=True, employee__department=dept
        ).count()
        if active_assets > 0:
            return Response(
                {"status": "error", "message": f"Cannot delete department with {active_assets} active asset(s)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        dept.delete()
        _org_audit(request, "department", key, dept.name, "deleted")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentDisableView(APIView):
    def post(self, request, key):
        try:
            dept = Department.objects.get(id=key, deleted=False)
        except Department.DoesNotExist:
            return Response({"status": "error", "message": "Department not found"}, status=status.HTTP_404_NOT_FOUND)
        prev_status = dept.status
        dept.status = "Disabled"
        dept.save(update_fields=["status"])
        _org_audit(request, "department", key, dept.name, "disabled",
                   prev={"status": prev_status}, new={"status": "Disabled"})
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentDashboardView(APIView):
    def get(self, request, key):
        try:
            dept = Department.objects.get(id=key, deleted=False)
        except Department.DoesNotExist:
            return Response({"status": "error", "message": "Department not found"}, status=status.HTTP_404_NOT_FOUND)
        employees = Employee.objects.filter(department=dept, deleted=False)
        assignments = EmployeeAssetAssignment.objects.filter(is_active=True, employee__department=dept)
        total_assets = assignments.values("client").distinct().count()
        online_devices = Client.objects.filter(
            employee_assignments__in=assignments, status="online", deleted=False
        ).distinct().count()
        offline_devices = total_assets - online_devices
        asset_value = sum(
            (a.client.purchase_cost for a in assignments.select_related("client") if a.client and a.client.purchase_cost),
            __import__("decimal").Decimal("0"),
        )
        return Response({
            "total_assets": total_assets,
            "assigned_assets": total_assets,
            "employees_count": employees.count(),
            "maintenance_count": Client.objects.filter(
                employee_assignments__in=assignments, status="offline", deleted=False
            ).distinct().count(),
            "asset_value": float(asset_value),
            "online_devices": online_devices,
            "offline_devices": max(0, offline_devices),
        })


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentExportView(APIView):
    def get(self, request):
        import csv
        from django.http import HttpResponse
        company = get_user_company(request)
        fmt = request.query_params.get("format", "csv")
        qs = Department.objects.filter(deleted=False)
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(models.Q(name__icontains=search) | models.Q(code__icontains=search))
        if fmt == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="departments.csv"'
            writer = csv.writer(response)
            writer.writerow(["Name", "Code", "Head", "Email", "Phone", "Budget", "Status"])
            for d in qs:
                writer.writerow([d.name, d.code, d.department_head, d.email,
                                 d.phone_number, d.budget or "", d.status])
            return response
        return Response(DepartmentListSerializer(qs, many=True).data)


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentImportView(APIView):
    def post(self, request):
        import csv, io
        f = request.FILES.get("file")
        if not f:
            return Response({"status": "error", "message": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = request.data.get("dry_run", "false") == "true"
        try:
            content = f.read().decode("utf-8")
        except UnicodeDecodeError:
            return Response({"status": "error", "message": "File must be UTF-8 encoded"}, status=status.HTTP_400_BAD_REQUEST)
        reader = csv.DictReader(io.StringIO(content))
        success_count = 0
        errors = []
        for i, row in enumerate(reader, start=2):
            name = row.get("name", "").strip()
            code = row.get("code", "").strip()
            if not name or not code:
                errors.append({"row": i, "message": "name and code are required"})
                continue
            if Department.objects.filter(name=name, deleted=False).exists():
                errors.append({"row": i, "message": f"Department '{name}' already exists"})
                continue
            if Department.objects.filter(code=code, deleted=False).exists():
                errors.append({"row": i, "message": f"Department code '{code}' already exists"})
                continue
            if not dry_run:
                Department.objects.create(
                    name=name, code=code, description=row.get("description", ""),
                    department_head=row.get("department_head", ""),
                    email=row.get("email", ""), phone_number=row.get("phone_number", ""),
                    budget=row.get("budget") or None,
                    status=row.get("status", "Active"),
                )
            success_count += 1
        return Response({"status": "ok", "success_count": success_count, "errors": errors, "dry_run": dry_run})


@method_decorator(csrf_exempt, name="dispatch")
class DepartmentBulkActionView(APIView):
    def post(self, request):
        action = request.data.get("action", "")
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"status": "error", "message": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        company = get_user_company(request)
        qs = Department.objects.filter(id__in=ids, deleted=False)
        if company:
            qs = qs.filter(company=company)
        count = 0
        if action == "disable":
            count = qs.update(status="Disabled")
        elif action == "archive":
            count = qs.update(status="Archived")
        elif action == "delete":
            for dept in qs:
                if not Employee.objects.filter(department=dept, deleted=False).exists():
                    active = EmployeeAssetAssignment.objects.filter(
                        is_active=True, employee__department=dept
                    ).count()
                    if active == 0:
                        dept.delete()
                        count += 1
        return Response({"status": "ok", "count": count})


# ── Employee Views ──────────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = Employee.objects.filter(deleted=False).select_related("department", "location")
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        dept = request.query_params.get("department", "").strip()
        loc = request.query_params.get("location", "").strip()
        s = request.query_params.get("status", "").strip()
        has_assets = request.query_params.get("has_assets", "").strip()
        if search:
            qs = qs.filter(
                models.Q(full_name__icontains=search) |
                models.Q(employee_code__icontains=search) |
                models.Q(email__icontains=search)
            )
        if dept:
            qs = qs.filter(department_id=dept)
        if loc:
            qs = qs.filter(location_id=loc)
        if s:
            qs = qs.filter(status=s)
        if has_assets == "true":
            qs = qs.filter(asset_assignments__is_active=True).distinct()
        elif has_assets == "false":
            qs = qs.exclude(asset_assignments__is_active=True)
        return Response(EmployeeListSerializer(qs, many=True).data)

    def post(self, request):
        company = get_user_company(request)
        serializer = EmployeeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if Employee.objects.filter(email=data["email"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "An employee with this email already exists"}, status=status.HTTP_400_BAD_REQUEST)
        if Employee.objects.filter(employee_code=data["employee_code"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "An employee with this code already exists"}, status=status.HTTP_400_BAD_REQUEST)
        if data.get("phone_number") and Employee.objects.filter(phone_number=data["phone_number"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "An employee with this phone number already exists"}, status=status.HTTP_400_BAD_REQUEST)
        emp = serializer.save(company=company)
        _org_audit(request, "employee", emp.id, emp.full_name, "created", new=serializer.data, company=company)
        return Response(EmployeeSerializer(emp).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeDetailView(APIView):
    def get(self, request, key):
        try:
            emp = Employee.objects.select_related("department", "location", "reports_to").get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(EmployeeSerializer(emp).data)

    def put(self, request, key):
        try:
            emp = Employee.objects.get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        prev = EmployeeSerializer(emp).data
        serializer = EmployeeSerializer(emp, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        new_email = data.get("email", emp.email)
        if Employee.objects.filter(email=new_email, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "An employee with this email already exists"}, status=status.HTTP_400_BAD_REQUEST)
        new_code = data.get("employee_code", emp.employee_code)
        if Employee.objects.filter(employee_code=new_code, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "An employee with this code already exists"}, status=status.HTTP_400_BAD_REQUEST)
        new_phone = data.get("phone_number", emp.phone_number)
        if new_phone and Employee.objects.filter(phone_number=new_phone, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "An employee with this phone number already exists"}, status=status.HTTP_400_BAD_REQUEST)
        updated = serializer.save()
        _org_audit(request, "employee", key, updated.full_name, "updated", prev=prev, new=serializer.data)
        return Response(EmployeeSerializer(updated).data)


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeDeleteView(APIView):
    def post(self, request, key):
        try:
            emp = Employee.objects.get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        if EmployeeAssetAssignment.objects.filter(employee=emp).exists():
            return Response(
                {"status": "error", "message": "Cannot delete employee with asset assignment history"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        emp.delete()
        _org_audit(request, "employee", key, emp.full_name, "deleted")
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeDeactivateView(APIView):
    def post(self, request, key):
        try:
            emp = Employee.objects.get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        new_status = request.data.get("status", "Inactive")
        valid = ["Inactive", "Resigned", "On Leave", "Terminated", "Retired"]
        if new_status not in valid:
            return Response({"status": "error", "message": f"Invalid status. Must be one of: {', '.join(valid)}"},
                            status=status.HTTP_400_BAD_REQUEST)
        prev_status = emp.status
        emp.status = new_status
        emp.save(update_fields=["status"])
        if new_status != "Active":
            EmployeeAssetAssignment.objects.filter(employee=emp, is_active=True).update(
                is_active=False, returned_at=timezone.now()
            )
        _org_audit(request, "employee", key, emp.full_name, "deactivated",
                   prev={"status": prev_status}, new={"status": new_status})
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeArchiveView(APIView):
    def post(self, request, key):
        try:
            emp = Employee.objects.get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        prev_status = emp.status
        emp.status = "Inactive"
        emp.save(update_fields=["status"])
        EmployeeAssetAssignment.objects.filter(employee=emp, is_active=True).update(
            is_active=False, returned_at=timezone.now()
        )
        _org_audit(request, "employee", key, emp.full_name, "archived",
                   prev={"status": prev_status}, new={"status": "Inactive"})
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeDashboardView(APIView):
    def get(self, request, key):
        try:
            emp = Employee.objects.get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        all_assignments = EmployeeAssetAssignment.objects.filter(employee=emp)
        current = all_assignments.filter(is_active=True)
        previous = all_assignments.filter(is_active=False)
        current_assets = current.values("client").distinct().count()
        previous_assets = previous.values("client").distinct().count()
        asset_value = sum(
            (a.client.purchase_cost for a in current.select_related("client") if a.client and a.client.purchase_cost),
            __import__("decimal").Decimal("0"),
        )
        last_assignment = all_assignments.order_by("-assigned_at").first()
        return Response({
            "total_assigned_assets": all_assignments.count(),
            "current_assets": current_assets,
            "previous_assets": previous_assets,
            "asset_history": EmployeeAssetAssignmentSerializer(all_assignments[:20], many=True).data,
            "last_assignment_date": last_assignment.assigned_at.isoformat() if last_assignment else None,
            "asset_value": float(asset_value),
        })


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeAssetHistoryView(APIView):
    def get(self, request, key):
        try:
            emp = Employee.objects.get(id=key, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        qs = EmployeeAssetAssignment.objects.filter(employee=emp).select_related("client")
        return Response(EmployeeAssetAssignmentSerializer(qs, many=True).data)


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeExportView(APIView):
    def get(self, request):
        import csv
        from django.http import HttpResponse
        company = get_user_company(request)
        fmt = request.query_params.get("format", "csv")
        qs = Employee.objects.filter(deleted=False).select_related("department", "location")
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(models.Q(full_name__icontains=search) | models.Q(employee_code__icontains=search))
        if fmt == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="employees.csv"'
            writer = csv.writer(response)
            writer.writerow(["Code", "Name", "Email", "Phone", "Department", "Designation",
                             "Location", "Status", "Joining Date", "Manager"])
            for e in qs:
                writer.writerow([e.employee_code, e.full_name, e.email, e.phone_number,
                                 e.department.name if e.department else "", e.designation,
                                 e.location.office_name if e.location else "", e.status,
                                 e.joining_date or "", e.manager_name])
            return response
        return Response(EmployeeListSerializer(qs, many=True).data)


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeImportView(APIView):
    def post(self, request):
        import csv, io
        f = request.FILES.get("file")
        if not f:
            return Response({"status": "error", "message": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = request.data.get("dry_run", "false") == "true"
        try:
            content = f.read().decode("utf-8")
        except UnicodeDecodeError:
            return Response({"status": "error", "message": "File must be UTF-8 encoded"}, status=status.HTTP_400_BAD_REQUEST)
        reader = csv.DictReader(io.StringIO(content))
        success_count = 0
        errors = []
        for i, row in enumerate(reader, start=2):
            employee_code = row.get("employee_code", "").strip()
            full_name = row.get("full_name", "").strip()
            email = row.get("email", "").strip()
            dept_name = row.get("department", "").strip()
            loc_name = row.get("location", "").strip()
            if not employee_code or not full_name or not email:
                errors.append({"row": i, "message": "employee_code, full_name, and email are required"})
                continue
            if Employee.objects.filter(email=email, deleted=False).exists():
                errors.append({"row": i, "message": f"Employee with email '{email}' already exists"})
                continue
            if Employee.objects.filter(employee_code=employee_code, deleted=False).exists():
                errors.append({"row": i, "message": f"Employee code '{employee_code}' already exists"})
                continue
            if not dry_run:
                dept = Department.objects.filter(name=dept_name, deleted=False).first() if dept_name else None
                loc = Location.objects.filter(office_name=loc_name, deleted=False).first() if loc_name else None
                if not dept:
                    errors.append({"row": i, "message": f"Department '{dept_name}' not found"})
                    continue
                if not loc:
                    errors.append({"row": i, "message": f"Location '{loc_name}' not found"})
                    continue
                Employee.objects.create(
                    employee_code=employee_code, full_name=full_name, email=email,
                    phone_number=row.get("phone_number", ""), department=dept,
                    designation=row.get("designation", ""), location=loc,
                    joining_date=row.get("joining_date") or None,
                    manager_name=row.get("manager_name", ""),
                    status=row.get("status", "Active"),
                )
            success_count += 1
        return Response({"status": "ok", "success_count": success_count, "errors": errors, "dry_run": dry_run})


@method_decorator(csrf_exempt, name="dispatch")
class EmployeeBulkActionView(APIView):
    def post(self, request):
        action = request.data.get("action", "")
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"status": "error", "message": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        company = get_user_company(request)
        qs = Employee.objects.filter(id__in=ids, deleted=False)
        if company:
            qs = qs.filter(company=company)
        count = 0
        if action == "deactivate":
            count = qs.update(status="Inactive")
            EmployeeAssetAssignment.objects.filter(employee__in=qs, is_active=True).update(
                is_active=False, returned_at=timezone.now()
            )
        elif action == "archive":
            EmployeeAssetAssignment.objects.filter(employee__in=qs, is_active=True).update(
                is_active=False, returned_at=timezone.now()
            )
            count = qs.delete()[0]
        return Response({"status": "ok", "count": count})


# ── Assignment Views ────────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssignmentListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = EmployeeAssetAssignment.objects.all().select_related("employee", "client")
        if company:
            qs = qs.filter(employee__company=company)
        emp = request.query_params.get("employee", "").strip()
        active = request.query_params.get("active", "").strip()
        if emp:
            qs = qs.filter(employee_id=emp)
        if active == "true":
            qs = qs.filter(is_active=True)
        elif active == "false":
            qs = qs.filter(is_active=False)
        return Response(EmployeeAssetAssignmentSerializer(qs[:100], many=True).data)

    def post(self, request):
        serializer = EmployeeAssetAssignmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        emp = data["employee"]
        client = data.get("client")
        if emp.status != "Active":
            return Response(
                {"status": "error", "message": "Cannot assign assets to inactive employees"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if client and EmployeeAssetAssignment.objects.filter(client=client, is_active=True).exists():
            return Response(
                {"status": "error", "message": "This asset is already assigned to another employee"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        assignment = serializer.save(
            assigned_by=request.user.username if request.user.is_authenticated else "system"
        )
        return Response(EmployeeAssetAssignmentSerializer(assignment).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssignmentReturnView(APIView):
    def post(self, request, key):
        try:
            assignment = EmployeeAssetAssignment.objects.get(id=key, is_active=True)
        except EmployeeAssetAssignment.DoesNotExist:
            return Response({"status": "error", "message": "Active assignment not found"}, status=status.HTTP_404_NOT_FOUND)
        assignment.is_active = False
        assignment.returned_at = timezone.now()
        assignment.notes = request.data.get("notes", assignment.notes)
        assignment.save(update_fields=["is_active", "returned_at", "notes"])
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AssignmentBulkView(APIView):
    def post(self, request):
        action = request.data.get("action", "")
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"status": "error", "message": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        if action == "return":
            count = EmployeeAssetAssignment.objects.filter(id__in=ids, is_active=True).update(
                is_active=False, returned_at=timezone.now()
            )
        return Response({"status": "ok", "count": count})


# ── Org Audit & Stats ───────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class OrgAuditLogView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = OrgAuditLog.objects.all()
        if company:
            qs = qs.filter(company=company)
        entity_type = request.query_params.get("entity_type", "").strip()
        entity_id = request.query_params.get("entity_id", "").strip()
        action = request.query_params.get("action", "").strip()
        if entity_type:
            qs = qs.filter(entity_type=entity_type)
        if entity_id:
            qs = qs.filter(entity_id=entity_id)
        if action:
            qs = qs.filter(action=action)
        offset = int(request.query_params.get("offset", 0))
        limit = int(request.query_params.get("limit", 50))
        total = qs.count()
        entries = qs[offset:offset + limit]
        return Response({
            "total": total,
            "entries": OrgAuditLogSerializer(entries, many=True).data,
        })


@method_decorator(csrf_exempt, name="dispatch")
class OrgDashboardStatsView(APIView):
    def get(self, request):
        company = get_user_company(request)
        base_emp = Employee.objects.filter(deleted=False)
        base_dept = Department.objects.filter(deleted=False)
        base_loc = Location.objects.filter(deleted=False)
        base_asn = EmployeeAssetAssignment.objects.filter(is_active=True)
        if company:
            base_emp = base_emp.filter(company=company)
            base_dept = base_dept.filter(company=company)
            base_loc = base_loc.filter(company=company)
        total_employees = base_emp.count()
        active_employees = base_emp.filter(status="Active").count()
        total_departments = base_dept.count()
        total_locations = base_loc.count()
        total_assets = base_asn.values("client").distinct().count()
        return Response({
            "total_employees": total_employees,
            "active_employees": active_employees,
            "total_departments": total_departments,
            "total_locations": total_locations,
            "total_assets_assigned": total_assets,
        })


# ═══════════════════════════════════════════════════════════════════════════════
# ASSET MANAGEMENT MODULE
# ═══════════════════════════════════════════════════════════════════════════════


def _record_asset_history(asset, action, prev=None, new=None, request=None, notes=""):
    AssetHistory.objects.create(
        asset=asset,
        action=action,
        previous_value=prev or {},
        new_value=new or {},
        performed_by=request.user.username if request and request.user.is_authenticated else "system",
        ip_address=_client_ip(request) if request else None,
        user_agent=request.META.get("HTTP_USER_AGENT", "") if request else "",
        notes=notes,
    )


# ── Asset Category Views ────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetCategoryListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = AssetCategory.objects.filter(is_active=True)
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(models.Q(name__icontains=search) | models.Q(code__icontains=search))
        return Response(AssetCategoryListSerializer(qs, many=True).data)

    def post(self, request):
        company = get_user_company(request)
        serializer = AssetCategorySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        if AssetCategory.objects.filter(name=data["name"], company=company).exists():
            return Response({"status": "error", "message": "Category name already exists"},
                            status=status.HTTP_400_BAD_REQUEST)
        if AssetCategory.objects.filter(code=data["code"], company=company).exists():
            return Response({"status": "error", "message": "Category code already exists"},
                            status=status.HTTP_400_BAD_REQUEST)
        cat = serializer.save(company=company)
        return Response(AssetCategorySerializer(cat).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssetCategoryDetailView(APIView):
    def get(self, request, key):
        try:
            cat = AssetCategory.objects.get(id=key)
        except AssetCategory.DoesNotExist:
            return Response({"status": "error", "message": "Category not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(AssetCategorySerializer(cat).data)

    def put(self, request, key):
        try:
            cat = AssetCategory.objects.get(id=key)
        except AssetCategory.DoesNotExist:
            return Response({"status": "error", "message": "Category not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = AssetCategorySerializer(cat, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        new_name = serializer.validated_data.get("name", cat.name)
        if AssetCategory.objects.filter(name=new_name).exclude(id=key).exists():
            return Response({"status": "error", "message": "Category name already exists"},
                            status=status.HTTP_400_BAD_REQUEST)
        updated = serializer.save()
        return Response(AssetCategorySerializer(updated).data)

    def delete(self, request, key):
        try:
            cat = AssetCategory.objects.get(id=key)
        except AssetCategory.DoesNotExist:
            return Response({"status": "error", "message": "Category not found"}, status=status.HTTP_404_NOT_FOUND)
        if cat.assets.filter(deleted=False).exists():
            return Response({"status": "error", "message": "Cannot delete category with active assets"},
                            status=status.HTTP_400_BAD_REQUEST)
        cat.is_active = False
        cat.save(update_fields=["is_active"])
        return Response({"status": "ok"})


# ── Asset Vendor Views ──────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetVendorListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = AssetVendor.objects.all()
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(models.Q(name__icontains=search) | models.Q(contact_person__icontains=search))
        return Response(AssetVendorSerializer(qs, many=True).data)

    def post(self, request):
        company = get_user_company(request)
        serializer = AssetVendorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vendor = serializer.save(company=company)
        return Response(AssetVendorSerializer(vendor).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssetVendorDetailView(APIView):
    def get(self, request, key):
        try:
            vendor = AssetVendor.objects.get(id=key)
        except AssetVendor.DoesNotExist:
            return Response({"status": "error", "message": "Vendor not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(AssetVendorSerializer(vendor).data)

    def put(self, request, key):
        try:
            vendor = AssetVendor.objects.get(id=key)
        except AssetVendor.DoesNotExist:
            return Response({"status": "error", "message": "Vendor not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = AssetVendorSerializer(vendor, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        updated = serializer.save()
        return Response(AssetVendorSerializer(updated).data)

    def delete(self, request, key):
        try:
            vendor = AssetVendor.objects.get(id=key)
        except AssetVendor.DoesNotExist:
            return Response({"status": "error", "message": "Vendor not found"}, status=status.HTTP_404_NOT_FOUND)
        if vendor.assets.filter(deleted=False).exists():
            return Response({"status": "error", "message": "Cannot delete vendor with active assets"},
                            status=status.HTTP_400_BAD_REQUEST)
        vendor.delete()
        return Response({"status": "ok"})


# ── Asset Views ─────────────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetListView(APIView):
    def get(self, request):
        company = get_user_company(request)
        qs = Asset.objects.filter(deleted=False).select_related(
            "category", "vendor", "department", "location", "assigned_to", "parent", "client"
        )
        if company:
            qs = qs.filter(company=company)
        search = request.query_params.get("search", "").strip()
        category = request.query_params.get("category", "").strip()
        department = request.query_params.get("department", "").strip()
        location = request.query_params.get("location", "").strip()
        s = request.query_params.get("status", "").strip()
        warranty = request.query_params.get("warranty", "").strip()
        assigned = request.query_params.get("assigned", "").strip()

        if search:
            qs = qs.filter(
                models.Q(asset_name__icontains=search) |
                models.Q(asset_id__icontains=search) |
                models.Q(asset_tag__icontains=search) |
                models.Q(serial_number__icontains=search) |
                models.Q(manufacturer__icontains=search) |
                models.Q(model_name__icontains=search)
            )
        if category:
            qs = qs.filter(category_id=category)
        if department:
            qs = qs.filter(department_id=department)
        if location:
            qs = qs.filter(location_id=location)
        if s:
            qs = qs.filter(asset_status=s)
        if assigned == "true":
            qs = qs.filter(assigned_to__isnull=False)
        elif assigned == "false":
            qs = qs.filter(assigned_to__isnull=True)

        if warranty == "expiring":
            from django.utils import timezone as tz
            today = tz.now().date()
            from datetime import timedelta
            soon = today + timedelta(days=30)
            qs = qs.filter(warranty_end__lte=soon, warranty_end__gte=today)
        elif warranty == "expired":
            from django.utils import timezone as tz
            qs = qs.filter(warranty_end__lt=tz.now().date())

        page_size = int(request.query_params.get("page_size", 100))
        page = int(request.query_params.get("page", 1))
        total = qs.count()
        start = (page - 1) * page_size
        assets = qs[start:start + page_size]

        return Response({
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size if page_size else 1,
            "results": AssetListSerializer(assets, many=True).data,
        })

    def post(self, request):
        company = get_user_company(request)
        serializer = AssetSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        if Asset.objects.filter(asset_tag=data["asset_tag"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "Asset tag already exists"},
                            status=status.HTTP_400_BAD_REQUEST)
        if Asset.objects.filter(serial_number=data["serial_number"], deleted=False, company=company).exists():
            return Response({"status": "error", "message": "Serial number already exists"},
                            status=status.HTTP_400_BAD_REQUEST)

        asset = serializer.save(
            created_by=request.user.username if request.user.is_authenticated else "system",
            company=company,
        )
        _record_asset_history(asset, "created", new=AssetSerializer(asset).data, request=request,
                              notes=f"Asset {asset.asset_name} created")
        return Response(AssetSerializer(asset).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssetDetailView(APIView):
    def get(self, request, key):
        try:
            asset = Asset.objects.select_related(
                "category", "vendor", "department", "location", "assigned_to", "parent", "client"
            ).get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(AssetSerializer(asset).data)

    def put(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        if asset.asset_status in ("Retired", "Disposed"):
            return Response({"status": "error", "message": "Retired or disposed assets are read-only"},
                            status=status.HTTP_400_BAD_REQUEST)

        prev = AssetSerializer(asset).data
        serializer = AssetSerializer(asset, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        new_tag = data.get("asset_tag", asset.asset_tag)
        if Asset.objects.filter(asset_tag=new_tag, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "Asset tag already exists"},
                            status=status.HTTP_400_BAD_REQUEST)
        new_serial = data.get("serial_number", asset.serial_number)
        if Asset.objects.filter(serial_number=new_serial, deleted=False).exclude(id=key).exists():
            return Response({"status": "error", "message": "Serial number already exists"},
                            status=status.HTTP_400_BAD_REQUEST)

        updated = serializer.save()
        new = AssetSerializer(updated).data
        _record_asset_history(updated, "updated", prev=prev, new=new, request=request,
                              notes="Asset details updated")
        return Response(AssetSerializer(updated).data)

    def delete(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)
        if asset.asset_status == "Assigned":
            return Response({"status": "error", "message": "Cannot delete an assigned asset. Return it first."},
                            status=status.HTTP_400_BAD_REQUEST)
        prev = AssetSerializer(asset).data
        _record_asset_history(asset, "disposed", prev=prev, new={"deleted": True}, request=request,
                              notes="Asset deleted")
        asset.delete()
        return Response({"status": "ok"})


# ── Asset Status Engine ─────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetStatusView(APIView):
    def post(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        new_status = request.data.get("status", "").strip()
        notes = request.data.get("notes", "")
        if not new_status:
            return Response({"status": "error", "message": "Status is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        valid, msg = Asset.validate_status_transition(asset.asset_status, new_status)
        if not valid:
            return Response({"status": "error", "message": msg}, status=status.HTTP_400_BAD_REQUEST)

        prev_status = asset.asset_status
        prev = AssetSerializer(asset).data
        asset.asset_status = new_status
        asset.save(update_fields=["asset_status", "updated_at"])

        action = "status_changed"
        if new_status == "Maintenance":
            action = "maintenance_started"
        elif prev_status == "Maintenance" and new_status == "Available":
            action = "maintenance_completed"
        elif new_status == "Retired":
            action = "retired"
        elif new_status == "Disposed":
            action = "disposed"

        _record_asset_history(asset, action, prev={"asset_status": prev_status},
                              new={"asset_status": new_status}, request=request, notes=notes)
        return Response(AssetSerializer(asset).data)


# ── Asset Assignment Views ──────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetAssignView(APIView):
    def post(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        employee_id = request.data.get("employee")
        notes = request.data.get("notes", "")
        expected_return = request.data.get("expected_return_date")

        if not employee_id:
            return Response({"status": "error", "message": "Employee ID is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Employee not found"},
                            status=status.HTTP_404_NOT_FOUND)

        if employee.status != "Active":
            return Response({"status": "error", "message": "Cannot assign assets to inactive employees"},
                            status=status.HTTP_400_BAD_REQUEST)

        if asset.asset_status not in ("Available", "Returned"):
            return Response({"status": "error",
                             "message": f"Asset must be Available or Returned to assign. Current status: {asset.asset_status}"},
                            status=status.HTTP_400_BAD_REQUEST)

        if AssetAssignment.objects.filter(asset=asset, is_active=True).exists():
            return Response({"status": "error", "message": "Asset is already assigned to another employee"},
                            status=status.HTTP_400_BAD_REQUEST)

        prev = AssetSerializer(asset).data
        assignment = AssetAssignment.objects.create(
            asset=asset,
            employee=employee,
            department=employee.department,
            location=employee.location,
            assigned_by=request.user.username if request.user.is_authenticated else "system",
            assignment_notes=notes,
            expected_return_date=expected_return or None,
        )
        asset.asset_status = "Assigned"
        asset.assigned_to = employee
        asset.department = employee.department
        asset.location = employee.location
        asset.save(update_fields=["asset_status", "assigned_to", "department", "location", "updated_at"])

        _record_asset_history(asset, "assigned",
                              prev={"asset_status": prev.get("asset_status"), "assigned_to": prev.get("assigned_to")},
                              new={"asset_status": "Assigned", "assigned_to": str(employee.id)},
                              request=request, notes=notes)
        return Response(AssetAssignmentSerializerV2(assignment).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssetReturnView(APIView):
    def post(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        if asset.asset_status != "Assigned":
            return Response({"status": "error", "message": "Asset is not currently assigned"},
                            status=status.HTTP_400_BAD_REQUEST)

        return_notes = request.data.get("notes", "")
        prev = AssetSerializer(asset).data

        assignment = AssetAssignment.objects.filter(asset=asset, is_active=True).first()
        if assignment:
            assignment.is_active = False
            assignment.returned_at = timezone.now()
            assignment.return_notes = return_notes
            assignment.save(update_fields=["is_active", "returned_at", "return_notes"])

        asset.asset_status = "Returned"
        asset.assigned_to = None
        asset.save(update_fields=["asset_status", "assigned_to", "updated_at"])

        _record_asset_history(asset, "returned",
                              prev={"asset_status": prev.get("asset_status")},
                              new={"asset_status": "Returned"},
                              request=request, notes=return_notes)
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AssetTransferView(APIView):
    def post(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        to_employee_id = request.data.get("to_employee")
        reason = request.data.get("reason", "")
        notes = request.data.get("notes", "")

        if not to_employee_id:
            return Response({"status": "error", "message": "Destination employee is required"},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            to_employee = Employee.objects.get(id=to_employee_id, deleted=False)
        except Employee.DoesNotExist:
            return Response({"status": "error", "message": "Destination employee not found"},
                            status=status.HTTP_404_NOT_FOUND)

        if to_employee.status != "Active":
            return Response({"status": "error", "message": "Cannot transfer to inactive employees"},
                            status=status.HTTP_400_BAD_REQUEST)

        from_employee = asset.assigned_to
        prev = AssetSerializer(asset).data

        transfer = AssetTransfer.objects.create(
            asset=asset,
            from_employee=from_employee,
            to_employee=to_employee,
            from_department=asset.department,
            to_department=to_employee.department,
            from_location=asset.location,
            to_location=to_employee.location,
            reason=reason,
            transferred_by=request.user.username if request.user.is_authenticated else "system",
            notes=notes,
        )

        old_assignment = AssetAssignment.objects.filter(asset=asset, is_active=True).first()
        if old_assignment:
            old_assignment.is_active = False
            old_assignment.returned_at = timezone.now()
            old_assignment.return_notes = f"Transferred to {to_employee.full_name}"
            old_assignment.save(update_fields=["is_active", "returned_at", "return_notes"])

        new_assignment = AssetAssignment.objects.create(
            asset=asset,
            employee=to_employee,
            department=to_employee.department,
            location=to_employee.location,
            assigned_by=request.user.username if request.user.is_authenticated else "system",
            assignment_notes=f"Transfer from {from_employee.full_name if from_employee else 'N/A'}: {reason}",
        )

        asset.assigned_to = to_employee
        asset.department = to_employee.department
        asset.location = to_employee.location
        asset.save(update_fields=["assigned_to", "department", "location", "updated_at"])

        _record_asset_history(asset, "transferred",
                              prev={"assigned_to": prev.get("assigned_to"), "department": prev.get("department"),
                                    "location": prev.get("location")},
                              new={"assigned_to": str(to_employee.id), "department": str(to_employee.department_id),
                                   "location": str(to_employee.location_id)},
                              request=request, notes=reason)
        return Response(AssetTransferSerializer(transfer).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssetRetireView(APIView):
    def post(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        valid, msg = Asset.validate_status_transition(asset.asset_status, "Retired")
        if not valid:
            return Response({"status": "error", "message": msg}, status=status.HTTP_400_BAD_REQUEST)

        notes = request.data.get("notes", "")
        prev = AssetSerializer(asset).data
        asset.asset_status = "Retired"
        asset.save(update_fields=["asset_status", "updated_at"])

        AssetAssignment.objects.filter(asset=asset, is_active=True).update(
            is_active=False, returned_at=timezone.now(), return_notes="Asset retired"
        )

        _record_asset_history(asset, "retired",
                              prev={"asset_status": prev.get("asset_status")},
                              new={"asset_status": "Retired"},
                              request=request, notes=notes)
        return Response({"status": "ok"})


@method_decorator(csrf_exempt, name="dispatch")
class AssetDisposeView(APIView):
    def post(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        valid, msg = Asset.validate_status_transition(asset.asset_status, "Disposed")
        if not valid:
            return Response({"status": "error", "message": msg}, status=status.HTTP_400_BAD_REQUEST)

        notes = request.data.get("notes", "")
        prev = AssetSerializer(asset).data
        asset.asset_status = "Disposed"
        asset.assigned_to = None
        asset.save(update_fields=["asset_status", "assigned_to", "updated_at"])

        _record_asset_history(asset, "disposed",
                              prev={"asset_status": prev.get("asset_status")},
                              new={"asset_status": "Disposed"},
                              request=request, notes=notes)
        return Response({"status": "ok"})


# ── Asset History View ──────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetHistoryView(APIView):
    def get(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        limit = int(request.query_params.get("limit", 100))
        offset = int(request.query_params.get("offset", 0))
        action = request.query_params.get("action", "").strip()

        qs = AssetHistory.objects.filter(asset=asset)
        if action:
            qs = qs.filter(action=action)
        total = qs.count()
        entries = qs[offset:offset + limit]
        return Response({
            "total": total,
            "entries": AssetHistorySerializer(entries, many=True).data,
        })


# ── Asset Assignment List View ──────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetAssignmentListView(APIView):
    def get(self, request):
        qs = AssetAssignment.objects.all().select_related("asset", "employee", "department", "location")
        active = request.query_params.get("active", "").strip()
        emp = request.query_params.get("employee", "").strip()
        asset_id = request.query_params.get("asset", "").strip()

        if active == "true":
            qs = qs.filter(is_active=True)
        elif active == "false":
            qs = qs.filter(is_active=False)
        if emp:
            qs = qs.filter(employee_id=emp)
        if asset_id:
            qs = qs.filter(asset_id=asset_id)

        limit = int(request.query_params.get("limit", 100))
        return Response(AssetAssignmentSerializerV2(qs[:limit], many=True).data)

    def post(self, request):
        serializer = AssetAssignmentSerializerV2(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        asset = data["asset"]
        employee = data["employee"]

        if employee.status != "Active":
            return Response({"status": "error", "message": "Cannot assign to inactive employees"},
                            status=status.HTTP_400_BAD_REQUEST)
        if asset.asset_status not in ("Available", "Returned"):
            return Response({"status": "error", "message": f"Asset must be Available or Returned. Current: {asset.asset_status}"},
                            status=status.HTTP_400_BAD_REQUEST)
        if AssetAssignment.objects.filter(asset=asset, is_active=True).exists():
            return Response({"status": "error", "message": "Asset is already assigned"},
                            status=status.HTTP_400_BAD_REQUEST)

        assignment = serializer.save(
            assigned_by=request.user.username if request.user.is_authenticated else "system"
        )
        asset.asset_status = "Assigned"
        asset.assigned_to = employee
        asset.department = employee.department
        asset.location = employee.location
        asset.save(update_fields=["asset_status", "assigned_to", "department", "location", "updated_at"])

        _record_asset_history(asset, "assigned",
                              new={"asset_status": "Assigned", "assigned_to": str(employee.id)},
                              request=request)
        return Response(AssetAssignmentSerializerV2(assignment).data, status=status.HTTP_201_CREATED)


@method_decorator(csrf_exempt, name="dispatch")
class AssetAssignmentReturnView(APIView):
    def post(self, request, key):
        try:
            assignment = AssetAssignment.objects.get(id=key, is_active=True)
        except AssetAssignment.DoesNotExist:
            return Response({"status": "error", "message": "Active assignment not found"},
                            status=status.HTTP_404_NOT_FOUND)

        return_notes = request.data.get("notes", "")
        prev_status = assignment.asset.asset_status

        assignment.is_active = False
        assignment.returned_at = timezone.now()
        assignment.return_notes = return_notes
        assignment.save(update_fields=["is_active", "returned_at", "return_notes"])

        asset = assignment.asset
        asset.asset_status = "Returned"
        asset.assigned_to = None
        asset.save(update_fields=["asset_status", "assigned_to", "updated_at"])

        _record_asset_history(asset, "returned",
                              prev={"asset_status": prev_status},
                              new={"asset_status": "Returned"},
                              request=request, notes=return_notes)
        return Response({"status": "ok"})


# ── Asset Transfer List View ────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetTransferListView(APIView):
    def get(self, request):
        qs = AssetTransfer.objects.all().select_related("asset", "from_employee", "to_employee")
        asset_id = request.query_params.get("asset", "").strip()
        if asset_id:
            qs = qs.filter(asset_id=asset_id)
        limit = int(request.query_params.get("limit", 100))
        return Response(AssetTransferSerializer(qs[:limit], many=True).data)


# ── Asset QR Code View ──────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetQRCodeView(APIView):
    def get(self, request, key):
        try:
            asset = Asset.objects.get(id=key, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            import qrcode
            import base64
            from io import BytesIO

            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(str(asset.qr_code))
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

            barcode_img = qrcode.QRCode(version=1, box_size=10, border=3)
            barcode_img.add_data(str(asset.barcode))
            barcode_img.make(fit=True)
            bc = barcode_img.make_image(fill_color="black", back_color="white")
            bc_buffer = BytesIO()
            bc.save(bc_buffer, format="PNG")
            bc_base64 = base64.b64encode(bc_buffer.getvalue()).decode("utf-8")

            return Response({
                "qr_code": f"data:image/png;base64,{img_base64}",
                "barcode": f"data:image/png;base64,{bc_base64}",
                "qr_value": str(asset.qr_code),
                "barcode_value": str(asset.barcode),
            })
        except ImportError:
            return Response({
                "qr_value": str(asset.qr_code),
                "barcode_value": str(asset.barcode),
                "message": "QR image generation not available (qrcode library not installed)",
            })


@method_decorator(csrf_exempt, name="dispatch")
class AssetQRScanView(APIView):
    def get(self, request, qr_code):
        try:
            asset = Asset.objects.select_related(
                "category", "vendor", "department", "location", "assigned_to"
            ).get(qr_code=qr_code, deleted=False)
        except Asset.DoesNotExist:
            return Response({"status": "error", "message": "Asset not found for this QR code"},
                            status=status.HTTP_404_NOT_FOUND)
        return Response(AssetSerializer(asset).data)


# ── Asset Bulk Actions ──────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetBulkActionView(APIView):
    def post(self, request):
        action = request.data.get("action", "")
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"status": "error", "message": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        company = get_user_company(request)
        qs = Asset.objects.filter(id__in=ids, deleted=False)
        if company:
            qs = qs.filter(company=company)
        count = 0
        if action == "delete":
            for asset in qs:
                if asset.asset_status != "Assigned":
                    asset.delete()
                    count += 1
        elif action == "retire":
            for asset in qs:
                valid, _ = Asset.validate_status_transition(asset.asset_status, "Retired")
                if valid:
                    asset.asset_status = "Retired"
                    asset.save(update_fields=["asset_status", "updated_at"])
                    AssetAssignment.objects.filter(asset=asset, is_active=True).update(
                        is_active=False, returned_at=timezone.now()
                    )
                    count += 1
        elif action == "deactivate":
            count = qs.update(is_active=False)
        return Response({"status": "ok", "count": count})


# ── Asset Import View ───────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetImportView(APIView):
    def post(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"status": "error", "message": "No file uploaded"},
                            status=status.HTTP_400_BAD_REQUEST)
        dry_run = request.data.get("dry_run", "false") == "true"

        try:
            if f.name.endswith(".xlsx"):
                import openpyxl
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                import csv, io
                content = f.read().decode("utf-8")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
        except Exception as e:
            return Response({"status": "error", "message": f"File parse error: {str(e)}"},
                            status=status.HTTP_400_BAD_REQUEST)

        success_count = 0
        errors = []
        for i, row in enumerate(rows, start=2):
            asset_name = (row.get("asset_name") or row.get("name") or "").strip()
            asset_tag = (row.get("asset_tag") or row.get("tag") or "").strip()
            serial_number = (row.get("serial_number") or row.get("serial") or "").strip()

            if not asset_name or not asset_tag or not serial_number:
                errors.append({"row": i, "message": "asset_name, asset_tag, and serial_number are required"})
                continue
            if Asset.objects.filter(asset_tag=asset_tag, deleted=False).exists():
                errors.append({"row": i, "message": f"Asset tag '{asset_tag}' already exists"})
                continue
            if Asset.objects.filter(serial_number=serial_number, deleted=False).exists():
                errors.append({"row": i, "message": f"Serial number '{serial_number}' already exists"})
                continue

            dept_name = (row.get("department") or "").strip()
            loc_name = (row.get("location") or "").strip()
            cat_name = (row.get("category") or "").strip()
            vendor_name = (row.get("vendor") or "").strip()

            if not dry_run:
                dept = Department.objects.filter(name=dept_name, deleted=False).first() if dept_name else None
                loc = Location.objects.filter(office_name=loc_name, deleted=False).first() if loc_name else None
                cat = AssetCategory.objects.filter(name=cat_name, is_active=True).first() if cat_name else None
                ven = AssetVendor.objects.filter(name=vendor_name).first() if vendor_name else None

                purchase_cost = None
                try:
                    val = row.get("purchase_cost") or row.get("cost")
                    if val:
                        purchase_cost = float(val)
                except (ValueError, TypeError):
                    pass

                asset = Asset(
                    asset_name=asset_name,
                    asset_tag=asset_tag,
                    serial_number=serial_number,
                    manufacturer=(row.get("manufacturer") or "").strip(),
                    model_name=(row.get("model") or row.get("model_name") or "").strip(),
                    description=(row.get("description") or "").strip(),
                    category=cat,
                    department=dept,
                    location=loc,
                    vendor=ven,
                    purchase_cost=purchase_cost,
                    asset_status=row.get("status") or "Draft",
                    created_by=request.user.username if request.user.is_authenticated else "system",
                )
                asset.save()
                _record_asset_history(asset, "created", new=AssetSerializer(asset).data, request=request,
                                      notes="Imported from file")
            success_count += 1

        return Response({"status": "ok", "success_count": success_count, "errors": errors, "dry_run": dry_run})


# ── Asset Export View ───────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetExportView(APIView):
    def get(self, request):
        import csv
        from django.http import HttpResponse

        company = get_user_company(request)
        fmt = request.query_params.get("format", "csv")
        qs = Asset.objects.filter(deleted=False).select_related("category", "vendor", "department", "location")
        if company:
            qs = qs.filter(company=company)

        search = request.query_params.get("search", "").strip()
        category = request.query_params.get("category", "").strip()
        department = request.query_params.get("department", "").strip()
        location = request.query_params.get("location", "").strip()
        s = request.query_params.get("status", "").strip()

        if search:
            qs = qs.filter(
                models.Q(asset_name__icontains=search) |
                models.Q(asset_tag__icontains=search) |
                models.Q(serial_number__icontains=search)
            )
        if category:
            qs = qs.filter(category_id=category)
        if department:
            qs = qs.filter(department_id=department)
        if location:
            qs = qs.filter(location_id=location)
        if s:
            qs = qs.filter(asset_status=s)

        if fmt == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="assets.csv"'
            writer = csv.writer(response)
            writer.writerow([
                "Asset ID", "Asset Name", "Asset Tag", "Serial Number", "Category",
                "Manufacturer", "Model", "Status", "Department", "Location",
                "Assigned To", "Purchase Date", "Purchase Cost", "Current Value",
                "Warranty End", "Vendor",
            ])
            for a in qs:
                writer.writerow([
                    a.asset_id, a.asset_name, a.asset_tag, a.serial_number,
                    a.category.name if a.category else "",
                    a.manufacturer, a.model_name, a.asset_status,
                    a.department.name if a.department else "",
                    a.location.office_name if a.location else "",
                    a.assigned_to.full_name if a.assigned_to else "",
                    a.purchase_date or "", a.purchase_cost or "",
                    a.current_value or "", a.warranty_end or "",
                    a.vendor.name if a.vendor else "",
                ])
            return response

        return Response(AssetListSerializer(qs, many=True).data)


# ── Asset Dashboard View ────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetDashboardView(APIView):
    def get(self, request):
        from django.db.models import Count, Sum, Avg, Q
        from django.utils import timezone as tz
        from datetime import timedelta

        company = get_user_company(request)
        base_qs = Asset.objects.filter(deleted=False)
        if company:
            base_qs = base_qs.filter(company=company)

        total = base_qs.count()
        by_status = dict(
            base_qs.values_list("asset_status")
            .annotate(c=Count("id")).values_list("asset_status", "c")
        )

        assigned = by_status.get("Assigned", 0)
        available = by_status.get("Available", 0) + by_status.get("Returned", 0)
        maintenance = by_status.get("Maintenance", 0)
        retired = by_status.get("Retired", 0)
        disposed = by_status.get("Disposed", 0)

        today = tz.now().date()
        warranty_expiring = base_qs.filter(
            warranty_end__lte=today + timedelta(days=30),
            warranty_end__gte=today,
        ).count()

        total_value = base_qs.aggregate(
            total=Sum("current_value"))["total"] or 0

        by_category = list(
            base_qs.exclude(category__isnull=True)
            .values(name=models.F("category__name"))
            .annotate(count=Count("id")).order_by("-count")[:10]
        )

        by_department = list(
            base_qs.exclude(department__isnull=True)
            .values(name=models.F("department__name"))
            .annotate(count=Count("id")).order_by("-count")[:10]
        )

        by_location = list(
            base_qs.exclude(location__isnull=True)
            .values(name=models.F("location__office_name"))
            .annotate(count=Count("id")).order_by("-count")[:10]
        )

        recent_assets = AssetListSerializer(
            base_qs.order_by("-created_at")[:10], many=True
        ).data

        return Response({
            "total_assets": total,
            "assigned_assets": assigned,
            "available_assets": available,
            "maintenance_assets": maintenance,
            "retired_assets": retired,
            "disposed_assets": disposed,
            "warranty_expiring": warranty_expiring,
            "total_value": float(total_value),
            "by_status": by_status,
            "by_category": by_category,
            "by_department": by_department,
            "by_location": by_location,
            "recent_assets": recent_assets,
        })


# ── Asset Analytics View ────────────────────────────────────────────────────


@method_decorator(csrf_exempt, name="dispatch")
class AssetAnalyticsView(APIView):
    def get(self, request):
        from django.db.models import Count, Sum, Avg, Q, F
        from django.utils import timezone as tz
        from datetime import timedelta

        company = get_user_company(request)
        base_qs = Asset.objects.filter(deleted=False)
        if company:
            base_qs = base_qs.filter(company=company)

        total = base_qs.count()
        assigned_count = base_qs.filter(asset_status="Assigned").count()
        utilization_rate = round((assigned_count / total * 100), 1) if total else 0

        avg_age = base_qs.filter(purchase_date__isnull=False).aggregate(
            avg=Avg(F("created_at__date") - F("purchase_date"))
        )["avg"]
        avg_age_days = None
        if avg_age:
            avg_age_days = avg_age.days if hasattr(avg_age, "days") else int(avg_age)

        value_by_department = list(
            base_qs.exclude(department__isnull=True)
            .values(name=models.F("department__name"))
            .annotate(total_value=Sum("current_value"), count=Count("id"))
            .order_by("-total_value")[:10]
        )

        lifecycle_dist = dict(
            base_qs.values_list("asset_status")
            .annotate(c=Count("id")).values_list("asset_status", "c")
        )

        today = tz.now().date()
        year_ago = today - timedelta(days=365)
        monthly_growth = []
        for i in range(12):
            month_start = (today - timedelta(days=30 * (11 - i))).replace(day=1)
            if i < 11:
                month_end = (today - timedelta(days=30 * (10 - i))).replace(day=1)
            else:
                month_end = today
            count = base_qs.filter(
                created_at__date__gte=month_start, created_at__date__lt=month_end
            ).count()
            monthly_growth.append({
                "month": month_start.strftime("%Y-%m"),
                "count": count,
            })

        warranty_expiring_30 = base_qs.filter(
            warranty_end__lte=today + timedelta(days=30),
            warranty_end__gte=today,
        ).count()

        warranty_expired = base_qs.filter(warranty_end__lt=today).count()

        depreciation_total = base_qs.filter(
            purchase_cost__isnull=False, current_value__isnull=False
        ).aggregate(
            total_depreciation=Sum(F("purchase_cost") - F("current_value"))
        )["total_depreciation"] or 0

        return Response({
            "utilization_rate": utilization_rate,
            "assigned_count": assigned_count,
            "total_assets": total,
            "avg_age_days": avg_age_days,
            "value_by_department": value_by_department,
            "lifecycle_distribution": lifecycle_dist,
            "monthly_growth": monthly_growth,
            "warranty_expiring_30_days": warranty_expiring_30,
            "warranty_expired": warranty_expired,
            "total_depreciation": float(depreciation_total),
        })


# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL SEARCH
# ═══════════════════════════════════════════════════════════════════════════════


@method_decorator(csrf_exempt, name="dispatch")
class GlobalSearchView(APIView):
    def get(self, request):
        q = request.query_params.get("q", "").strip()
        if not q or len(q) < 2:
            return Response({"results": []})

        from django.db.models import Q as QQ
        company = get_user_company(request)
        results = []

        asset_qs = Asset.objects.filter(deleted=False)
        emp_qs = Employee.objects.filter(deleted=False)
        dept_qs = Department.objects.filter(deleted=False)
        loc_qs = Location.objects.filter(deleted=False)
        client_qs = Client.objects.filter(deleted=False)
        if company:
            asset_qs = asset_qs.filter(company=company)
            emp_qs = emp_qs.filter(company=company)
            dept_qs = dept_qs.filter(company=company)
            loc_qs = loc_qs.filter(company=company)
            client_qs = client_qs.filter(company=company)

        # Assets
        for asset in asset_qs.filter(
            QQ(asset_name__icontains=q) | QQ(asset_tag__icontains=q) | QQ(serial_number__icontains=q) | QQ(model_name__icontains=q)
        )[:5]:
            results.append({
                "type": "asset", "id": str(asset.id),
                "title": asset.asset_name, "subtitle": f"{asset.asset_tag} | {asset.model_name or ''}",
                "url": f"/assets/{asset.id}/",
            })

        # Employees
        for emp in emp_qs.filter(
            QQ(full_name__icontains=q) | QQ(employee_code__icontains=q) | QQ(email__icontains=q)
        )[:5]:
            results.append({
                "type": "employee", "id": str(emp.id),
                "title": emp.full_name,
                "subtitle": emp.employee_code or emp.email or "",
                "url": "#",
            })

        # Departments
        for dep in dept_qs.filter(
            QQ(name__icontains=q) | QQ(code__icontains=q)
        )[:5]:
            results.append({
                "type": "department", "id": str(dep.id),
                "title": dep.name, "subtitle": dep.code or "",
                "url": "#",
            })

        # Locations
        for loc in loc_qs.filter(
            QQ(office_name__icontains=q) | QQ(city__icontains=q) | QQ(country__icontains=q)
        )[:5]:
            results.append({
                "type": "location", "id": str(loc.id),
                "title": loc.office_name, "subtitle": f"{loc.city or ''}, {loc.country or ''}",
                "url": "#",
            })

        # Clients
        for c in client_qs.filter(
            QQ(hostname__icontains=q) | QQ(registration_key__icontains=q) | QQ(platform__icontains=q)
        )[:5]:
            results.append({
                "type": "client", "id": c.registration_key,
                "title": c.hostname or "Unknown", "subtitle": c.registration_key,
                "url": f"/client/{c.registration_key}/",
            })

        # Licenses
        from maintenance.models import SoftwareLicense
        for lic in SoftwareLicense.objects.filter(deleted=False).filter(
            QQ(software_name__icontains=q) | QQ(license_key_masked__icontains=q)
        )[:5]:
            results.append({
                "type": "license", "id": str(lic.id),
                "title": lic.software_name, "subtitle": lic.license_key_masked or "",
                "url": "#",
            })

        # Alerts
        from intelligence.models import Alert
        for a in Alert.objects.filter(QQ(title__icontains=q) | QQ(description__icontains=q))[:5]:
            results.append({
                "type": "alert", "id": str(a.id),
                "title": a.title, "subtitle": f"{a.severity} | {a.status}",
                "url": "#",
            })

        return Response({"results": results, "total": len(results)})


# ═══════════════════════════════════════════════════════════════════════════════
# EXECUTIVE DASHBOARD — AGGREGATED ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════


@method_decorator(csrf_exempt, name="dispatch")
class ExecutiveAnalyticsView(APIView):
    """Aggregate KPIs from all modules for the executive dashboard."""

    def get(self, request):
        from django.db.models import Count, Sum, Avg, F
        from django.utils import timezone
        from datetime import timedelta, date

        today = timezone.now().date()
        User = get_user_model()
        company = get_user_company(request)

        from monitoring.models import DeviceMonitoringInfo, DeviceAlert
        from maintenance.models import MaintenanceRecord, SoftwareLicense
        from intelligence.models import Alert, AuditLogEntry

        # ── Base querysets (company-scoped) ──
        asset_qs = Asset.objects.filter(deleted=False)
        dept_qs = Department.objects.filter(deleted=False)
        emp_qs = Employee.objects.filter(deleted=False)
        loc_qs = Location.objects.filter(deleted=False)
        mon_qs = DeviceMonitoringInfo.objects.filter(client__deleted=False)
        mnt_qs = MaintenanceRecord.objects.filter(deleted=False)
        lic_qs = SoftwareLicense.objects.filter(deleted=False)
        alert_qs = Alert.objects.all()
        audit_qs = AuditLogEntry.objects.all()
        client_qs = Client.objects.filter(deleted=False)
        if company:
            asset_qs = asset_qs.filter(company=company)
            dept_qs = dept_qs.filter(company=company)
            emp_qs = emp_qs.filter(company=company)
            loc_qs = loc_qs.filter(company=company)
            mon_qs = mon_qs.filter(client__company=company)
            mnt_qs = mnt_qs.filter(company=company)
            lic_qs = lic_qs.filter(company=company)
            alert_qs = alert_qs.filter(client__company=company) if hasattr(Alert, 'client') else alert_qs
            audit_qs = audit_qs.filter(company=company) if hasattr(AuditLogEntry, 'company') else audit_qs
            client_qs = client_qs.filter(company=company)

        # ── Asset KPIs ──
        by_status = dict(
            asset_qs.values_list("asset_status")
            .annotate(c=Count("id")).values_list("asset_status", "c")
        )
        total_assets = sum(by_status.values())
        assigned_assets = by_status.get("Assigned", 0)
        available_assets = by_status.get("Available", 0) + by_status.get("Returned", 0)
        maintenance_assets = by_status.get("Maintenance", 0)
        retired_assets = by_status.get("Retired", 0)
        disposed_assets = by_status.get("Disposed", 0)
        warranty_expiring = asset_qs.filter(
            warranty_end__lte=today + timedelta(days=30), warranty_end__gte=today
        ).count()
        total_asset_value = asset_qs.aggregate(
            total=Sum("current_value"))["total"] or 0

        # ── Monitoring KPIs ──
        online_devices = mon_qs.filter(monitoring_status="online").count()
        offline_devices = mon_qs.filter(monitoring_status="offline").count()
        not_reporting = mon_qs.filter(monitoring_status="pending").count()
        critical_devices = DeviceAlert.objects.filter(status="active", severity="critical")
        if company:
            critical_devices = critical_devices.filter(client__company=company)
        critical_devices = critical_devices.count()

        # ── Maintenance KPIs ──
        upcoming_mnt = mnt_qs.filter(
            scheduled_date__lte=today + timedelta(days=30),
            scheduled_date__gte=today, status__in=("Approved", "Scheduled")
        ).count()
        overdue_mnt = mnt_qs.filter(
            status__in=("Scheduled", "In Progress", "Waiting Parts"), due_date__lt=today
        ).count()
        assets_under_repair = mnt_qs.filter(
            status__in=("In Progress", "Waiting Parts")
        ).values("asset").distinct().count()
        mnt_cost_ytd = mnt_qs.filter(
            completion_date__year=today.year, actual_cost__isnull=False
        ).aggregate(total=Sum("actual_cost"))["total"] or 0

        # ── License KPIs ──
        total_licenses = lic_qs.count()
        expiring_licenses = lic_qs.filter(
            expiration_date__lte=today + timedelta(days=30),
            expiration_date__gte=today
        ).count()
        expired_licenses = lic_qs.filter(status="Expired").count()
        used_seats = lic_qs.aggregate(total=Sum("seats_used"))["total"] or 0
        total_seats = lic_qs.aggregate(total=Sum("purchased_seats"))["total"] or 0
        license_compliance = round((used_seats / total_seats * 100), 1) if total_seats else 100

        # ── Security KPIs ──
        open_alerts = alert_qs.filter(status__in=("open", "acknowledged")).count()
        critical_alerts = alert_qs.filter(severity="critical", status="open").count()
        security_violations = alert_qs.filter(category__icontains="security", status="open").count()
        audit_today = audit_qs.filter(timestamp__date=today).count()

        # ── Charts: Asset Distribution ──
        asset_by_category = list(
            asset_qs.exclude(category__isnull=True)
            .values(name=F("category__name"))
            .annotate(count=Count("id")).order_by("-count")[:10]
        )
        asset_by_department = list(
            asset_qs.exclude(department__isnull=True)
            .values(name=F("department__name"))
            .annotate(count=Count("id")).order_by("-count")[:10]
        )
        lifecycle_dist = dict(
            asset_qs.values_list("asset_status")
            .annotate(c=Count("id")).values_list("asset_status", "c")
        )

        # ── Charts: Asset Growth Trend (12 months) ──
        growth_trend = []
        for i in range(12):
            ms = (today - timedelta(days=30 * (11 - i))).replace(day=1)
            me = (today - timedelta(days=30 * (10 - i))).replace(day=1) if i < 11 else today
            c = asset_qs.filter(created_at__date__gte=ms, created_at__date__lt=me).count()
            growth_trend.append({"month": ms.strftime("%Y-%m"), "count": c})

        # ── Charts: Monitoring ──
        monitor_status_dist = dict(
            mon_qs.values_list("monitoring_status").annotate(c=Count("id")).values_list("monitoring_status", "c")
        )
        avg_health = mon_qs.exclude(health_score=0).aggregate(avg=Avg("health_score"))["avg"] or 0

        # ── Charts: Maintenance Trends ──
        mnt_by_type = dict(
            mnt_qs.values_list("maintenance_type")
            .annotate(c=Count("id")).values_list("maintenance_type", "c")
        )

        # ── Charts: License Utilization ──
        license_by_type = dict(
            lic_qs.values_list("license_type")
            .annotate(c=Count("id")).values_list("license_type", "c")
        )

        # ── Charts: Security Trend (7 days) ──
        security_trend = []
        for i in range(7):
            d = today - timedelta(days=6 - i)
            c = alert_qs.filter(generated_time__date=d).count()
            security_trend.append({"date": d.isoformat(), "count": c})

        # ── Recent Activities ──
        recent_activities = []
        recent_activities += [
            {"type": "alert", "text": a.title, "time": a.generated_time.isoformat(), "severity": a.severity}
            for a in alert_qs.order_by("-generated_time")[:5]
        ]
        recent_activities += [
            {"type": "audit", "text": e.description, "time": e.timestamp.isoformat(), "username": e.username}
            for e in audit_qs.order_by("-timestamp")[:5]
        ]
        recent_activities = sorted(recent_activities, key=lambda x: x["time"], reverse=True)[:10]

        # ── Org Stats ──
        total_employees = emp_qs.count()
        total_departments = dept_qs.count()
        total_locations = loc_qs.count()
        total_users = User.objects.count()

        return Response({
            "asset_kpis": {
                "total_assets": total_assets,
                "assigned_assets": assigned_assets,
                "available_assets": available_assets,
                "maintenance_assets": maintenance_assets,
                "retired_assets": retired_assets,
                "disposed_assets": disposed_assets,
                "warranty_expiring": warranty_expiring,
                "total_value": float(total_asset_value),
            },
            "monitoring_kpis": {
                "online_devices": online_devices,
                "offline_devices": offline_devices,
                "not_reporting": not_reporting,
                "critical_devices": critical_devices,
                "avg_health_score": round(avg_health, 1),
                "status_distribution": monitor_status_dist,
            },
            "maintenance_kpis": {
                "upcoming_maintenance": upcoming_mnt,
                "overdue_maintenance": overdue_mnt,
                "assets_under_repair": assets_under_repair,
                "cost_ytd": float(mnt_cost_ytd),
                "by_type": mnt_by_type,
            },
            "license_kpis": {
                "total_licenses": total_licenses,
                "expiring_licenses": expiring_licenses,
                "expired_licenses": expired_licenses,
                "compliance_score": license_compliance,
                "by_type": license_by_type,
            },
            "security_kpis": {
                "open_alerts": open_alerts,
                "critical_alerts": critical_alerts,
                "security_violations": security_violations,
                "audit_events_today": audit_today,
                "trend": security_trend,
            },
            "charts": {
                "asset_by_category": asset_by_category,
                "asset_by_department": asset_by_department,
                "lifecycle_distribution": lifecycle_dist,
                "asset_growth_trend": growth_trend,
            },
            "organization": {
                "total_employees": total_employees,
                "total_departments": total_departments,
                "total_locations": total_locations,
                "total_users": total_users,
            },
            "recent_activities": recent_activities,
        })
